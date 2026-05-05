#!/usr/bin/env python
"""Check if the 5 problem sites were originally Techwise or SAS."""
import sys
sys.path.insert(0, 'src')

from openpyxl import load_workbook
from siteowlqa.survey_routing import DEFAULT_WORKBOOK_PATH

print("=" * 60)
print("CHECKING ORIGINAL VENDOR ASSIGNMENT")
print("=" * 60)

problem_sites = ['1590', '2646', '3072', '864', '9']

# Load the Survey Lab Excel to check original vendor
wb = load_workbook(DEFAULT_WORKBOOK_PATH, data_only=True)

# Check the Schedule tab
print("\n[1] Checking Schedule tab...")
if 'Schedule' in wb.sheetnames:
    ws = wb['Schedule']
    headers = [c.value for c in ws[1]]
    site_col = headers.index('Site') if 'Site' in headers else None
    vendor_col = headers.index('Vendor') if 'Vendor' in headers else None
    
    print(f"  Site column: {site_col}, Vendor column: {vendor_col}")
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        site = str(row[site_col]) if site_col is not None and row[site_col] else None
        vendor = row[vendor_col] if vendor_col is not None else None
        if site in problem_sites:
            print(f"  Site {site}: Vendor = {vendor}")

# Also check the Vendor ASSIGN tab for original assignment
print("\n[2] Checking Vendor ASSIGN tab...")
if 'Vendor ASSIGN' in wb.sheetnames:
    ws = wb['Vendor ASSIGN']
    headers = [c.value for c in ws[1]]
    
    # Find relevant columns
    site_col = None
    vendor_col = None
    state_col = None
    
    for i, h in enumerate(headers):
        if h and 'site' in str(h).lower():
            site_col = i
        if h and 'vendor' in str(h).lower() and 'survey' not in str(h).lower():
            vendor_col = i
        if h and 'state' in str(h).lower():
            state_col = i
    
    print(f"  Site col: {site_col}, Vendor col: {vendor_col}, State col: {state_col}")
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        site = str(row[site_col]).strip() if site_col is not None and row[site_col] else None
        if site in problem_sites:
            vendor = row[vendor_col] if vendor_col is not None else None
            state = row[state_col] if state_col is not None else None
            print(f"  Site {site}: Vendor = {vendor}, State = {state}")

wb.close()

# Now check ALL Hawaii and Alaska sites
print("\n[3] Checking Hawaii (HI) and Alaska (AK) sites...")
wb = load_workbook(DEFAULT_WORKBOOK_PATH, data_only=True)
ws = wb['Vendor ASSIGN']
headers = [c.value for c in ws[1]]

state_col = None
site_col = None
vendor_col = None
for i, h in enumerate(headers):
    if h and 'state' in str(h).lower():
        state_col = i
    if h and 'site' in str(h).lower():
        site_col = i
    if h and 'vendor' in str(h).lower() and 'survey' not in str(h).lower():
        vendor_col = i

hi_ak_sites = []
for row in ws.iter_rows(min_row=2, values_only=True):
    state = row[state_col] if state_col is not None else None
    if state in ['HI', 'AK']:
        site = str(row[site_col]).strip() if site_col is not None and row[site_col] else None
        vendor = row[vendor_col] if vendor_col is not None else None
        hi_ak_sites.append({'site': site, 'state': state, 'vendor': vendor})

print(f"  Found {len(hi_ak_sites)} Hawaii/Alaska sites:")
for s in hi_ak_sites:
    print(f"    Site {s['site']}: {s['state']} - Vendor: {s['vendor']}")

wb.close()
