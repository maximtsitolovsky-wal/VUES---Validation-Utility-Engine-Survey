#!/usr/bin/env python3
"""Count Excel sites."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from pathlib import Path

reference_path = Path(r"C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\ScoutSurveyLab.xlsm")
print(f"Loading: {reference_path}")
print(f"Exists: {reference_path.exists()}")

wb = openpyxl.load_workbook(reference_path, read_only=True, data_only=True)
print(f"Sheets: {wb.sheetnames}")

ws = wb["Scout Map Data"]
excel_sites = []
for row in ws.iter_rows(min_row=2, max_col=1):
    if row[0].value:
        excel_sites.append(str(row[0].value).strip().lstrip('0'))
wb.close()

print(f"Excel rows: {len(excel_sites)}")
print(f"Excel unique sites: {len(set(excel_sites))}")
