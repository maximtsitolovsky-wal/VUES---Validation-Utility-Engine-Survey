"""
Generate CEI Survey Report based on decision tree.
Pulls all CEI assignments and applies decision tree logic to determine survey type needed.
"""
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add src to path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from siteowlqa.vendor_assignment_tracker import VendorAssignmentTracker
from siteowlqa.survey_routing import evaluate_site, ScoutAnswers, ScheduleData, load_schedule_data
from siteowlqa.airtable_client import AirtableClient
from siteowlqa.config import get_config

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def get_cctv_triggers(scout: ScoutAnswers) -> List[str]:
    """Get CCTV survey triggers that are active."""
    triggers = []
    if scout.analog_baluns_present:
        triggers.append("Analog CCTV baluns present")
    if scout.rooftop_trimount_present:
        triggers.append("Rooftop tri-mount present")
    if scout.rooftop_trimount_count > 0:
        triggers.append(f"Rooftop tri-mount count: {scout.rooftop_trimount_count}")
    if scout.cable_condition:
        triggers.append(f"Cable condition: {scout.cable_condition}")
    if scout.homerun_cabling_present:
        triggers.append("Homerun cabling present")
    return triggers


def get_fa_triggers(scout: ScoutAnswers) -> List[str]:
    """Get FA/Intrusion survey triggers that are active."""
    triggers = []
    if scout.ceiling_mounted_devices:
        triggers.append("Ceiling mounted notification devices")
    if scout.sales_floor_column_devices:
        triggers.append("Sales floor column notification devices")
    if scout.emergency_exit_only_devices:
        triggers.append("Emergency exit only notification devices")
    if scout.fire_panel_type:
        triggers.append(f"Fire panel: {scout.fire_panel_type}")
    return triggers


def get_upgrade_triggers(scout: ScoutAnswers) -> Dict[str, bool]:
    """Determine upgrade triggers."""
    return {
        "fa_full_upgrade": scout.one_notification_device,
        "cctv_full_upgrade": scout.coax_siamese_cable,
    }


def main():
    config = get_config()
    
    # Load assignment tracker
    assignment_file = Path(config.vendor_assignment_file)
    if not assignment_file.exists():
        logger.error(f"Assignment file not found: {assignment_file}")
        return
    
    tracker = VendorAssignmentTracker(str(assignment_file))
    tracker.load_assignments()
    
    # Get CEI assignments
    cei_assignments = [a for a in tracker.assignments if a.vendor_name == "CEI"]
    logger.info(f"Found {len(cei_assignments)} CEI assignments")
    
    # Load schedule data (for Excel tracking info)
    schedule_data = load_schedule_data(config.workbook_path)
    schedule_by_site = {s.site: s for s in schedule_data}
    
    # Get Airtable client and fetch scout records
    airtable = AirtableClient(config)
    try:
        scout_records = airtable.fetch_all_records(
            config.scout_base_id,
            config.scout_table_id,
            batch_size=100
        )
    except Exception as e:
        logger.error(f"Failed to fetch scout records: {e}")
        scout_records = []
    
    # Build scout data map by site
    scout_by_site = {}
    for record in scout_records:
        site = str(record.get("fields", {}).get("Store Number", "")).strip()
        if site:
            raw_fields = record.get("fields", {})
            scout_by_site[site] = ScoutAnswers(
                site=site,
                record_id=record.get("id", ""),
                one_notification_device=raw_fields.get("One notification device") == True,
                ceiling_mounted_devices=raw_fields.get("Ceiling mounted devices") == True,
                sales_floor_column_devices=raw_fields.get("Sales floor column notification devices") == True,
                emergency_exit_only_devices=raw_fields.get("Emergency exit only notification devices") == True,
                fire_panel_type=str(raw_fields.get("Fire Panel Type", "")).strip(),
                coax_siamese_cable=raw_fields.get("Coax or Siamese Cable") == True,
                analog_baluns_present=raw_fields.get("Analog CCTV Baluns Present?") == True,
                rooftop_trimount_present=raw_fields.get("Rooftop Tri-Mount Present") == True,
                rooftop_trimount_count=int(raw_fields.get("Rooftop Tri-Mount Count", 0) or 0),
                cable_condition=str(raw_fields.get("Cable Condition", "")).strip(),
                homerun_cabling_present=raw_fields.get("Homerun Cabling Present") == True,
                ap_office_moving=str(raw_fields.get("AP office moving", "")).strip(),
                raw_fields=raw_fields
            )
    
    logger.info(f"Loaded {len(scout_by_site)} scout records")
    
    # Build report rows
    rows = []
    for assignment in sorted(cei_assignments, key=lambda a: int(a.site_number)):
        site = assignment.site_number
        scout = scout_by_site.get(site)
        schedule = schedule_by_site.get(site)
        
        # Apply decision tree
        routing = evaluate_site(scout, schedule)
        
        # Get triggers
        cctv_triggers = get_cctv_triggers(scout) if scout else []
        fa_triggers = get_fa_triggers(scout) if scout else []
        upgrades = get_upgrade_triggers(scout) if scout else {}
        
        row = {
            "site_number": site,
            "scout_submitted": "YES" if scout else "NO",
            "survey_required": routing.survey_required,
            "survey_type": routing.survey_type,
            "upgrade_decision": routing.upgrade_decision,
            "reason_for_decision": routing.reason_for_decision,
            "fa_full_upgrade": "YES" if upgrades.get("fa_full_upgrade") else "NO",
            "cctv_full_upgrade": "YES" if upgrades.get("cctv_full_upgrade") else "NO",
            "fa_triggers": "; ".join(fa_triggers) if fa_triggers else "—",
            "cctv_triggers": "; ".join(cctv_triggers) if cctv_triggers else "—",
            "supplemental_flags": routing.supplemental_flags or "—",
            "schedule_status": routing.schedule_status,
            "days_to_construction": routing.days_to_construction or "—",
            "ready_to_assign": routing.ready_to_assign,
            "assigned_vendor": routing.assigned_vendor or "—",
            "survey_complete": "YES" if routing.survey_complete else "NO",
            "percentage_complete": f"{routing.percentage_complete:.0%}" if routing.percentage_complete else "—",
        }
        rows.append(row)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CEI Surveys"
    
    # Define styles
    header_fill = PatternFill(start_color="0053E2", end_color="0053E2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Color code survey types
    survey_colors = {
        "BOTH": "FFC857",      # Spark yellow
        "CCTV": "5AE4FF",      # CEI cyan
        "FA/INTRUSION": "2a8703",  # Green
        "NONE": "95989A",      # Gray
        "PENDING": "FFC857",   # Warning yellow
        "REVIEW": "FFC857"     # Warning yellow
    }
    
    schedule_colors = {
        "URGENT": "EA1100",    # Red
        "ON TRACK": "2a8703",  # Green
        "COMPLETE": "2a8703", # Green
        "NOT REQUIRED": "95989A",  # Gray
        "PENDING": "FFC857",   # Yellow
        "REVIEW": "FFC857"     # Yellow
    }
    
    # Headers
    headers = [
        "Site Number",
        "Scout Submitted",
        "Survey Required",
        "Survey Type",
        "Upgrade Decision",
        "Reason for Decision",
        "FA/Intrusion Full Upgrade",
        "CCTV Full Upgrade",
        "FA/Intrusion Triggers",
        "CCTV Triggers",
        "Supplemental Flags",
        "Schedule Status",
        "Days to Construction",
        "Ready to Assign",
        "Assigned Vendor",
        "Survey Complete",
        "% Complete"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row_data["site_number"])
        ws.cell(row=row_idx, column=2, value=row_data["scout_submitted"])
        ws.cell(row=row_idx, column=3, value=row_data["survey_required"])
        ws.cell(row=row_idx, column=4, value=row_data["survey_type"])
        ws.cell(row=row_idx, column=5, value=row_data["upgrade_decision"])
        ws.cell(row=row_idx, column=6, value=row_data["reason_for_decision"])
        ws.cell(row=row_idx, column=7, value=row_data["fa_full_upgrade"])
        ws.cell(row=row_idx, column=8, value=row_data["cctv_full_upgrade"])
        ws.cell(row=row_idx, column=9, value=row_data["fa_triggers"])
        ws.cell(row=row_idx, column=10, value=row_data["cctv_triggers"])
        ws.cell(row=row_idx, column=11, value=row_data["supplemental_flags"])
        ws.cell(row=row_idx, column=12, value=row_data["schedule_status"])
        ws.cell(row=row_idx, column=13, value=row_data["days_to_construction"])
        ws.cell(row=row_idx, column=14, value=row_data["ready_to_assign"])
        ws.cell(row=row_idx, column=15, value=row_data["assigned_vendor"])
        ws.cell(row=row_idx, column=16, value=row_data["survey_complete"])
        ws.cell(row=row_idx, column=17, value=row_data["percentage_complete"])
        
        # Apply color coding
        survey_type = row_data["survey_type"]
        if survey_type in survey_colors:
            color = survey_colors[survey_type]
            ws.cell(row=row_idx, column=4).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        schedule_status = row_data["schedule_status"]
        if schedule_status in schedule_colors:
            color = schedule_colors[schedule_status]
            ws.cell(row=row_idx, column=12).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Apply borders and alignment
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12  # Site Number
    ws.column_dimensions['B'].width = 15  # Scout Submitted
    ws.column_dimensions['C'].width = 15  # Survey Required
    ws.column_dimensions['D'].width = 15  # Survey Type
    ws.column_dimensions['E'].width = 25  # Upgrade Decision
    ws.column_dimensions['F'].width = 35  # Reason
    ws.column_dimensions['G'].width = 15  # FA Full Upgrade
    ws.column_dimensions['H'].width = 15  # CCTV Full Upgrade
    ws.column_dimensions['I'].width = 30  # FA Triggers
    ws.column_dimensions['J'].width = 30  # CCTV Triggers
    ws.column_dimensions['K'].width = 35  # Supplemental Flags
    ws.column_dimensions['L'].width = 15  # Schedule Status
    ws.column_dimensions['M'].width = 15  # Days to Construction
    ws.column_dimensions['N'].width = 15  # Ready to Assign
    ws.column_dimensions['O'].width = 15  # Assigned Vendor
    ws.column_dimensions['P'].width = 15  # Survey Complete
    ws.column_dimensions['Q'].width = 12  # % Complete
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    # Calculate summary stats
    total_sites = len(rows)
    surveys_required = sum(1 for r in rows if r["survey_required"] == "YES")
    no_survey_needed = sum(1 for r in rows if r["survey_required"] == "NO")
    pending = sum(1 for r in rows if r["survey_required"] == "PENDING")
    
    survey_type_counts = {}
    for r in rows:
        st = r["survey_type"]
        survey_type_counts[st] = survey_type_counts.get(st, 0) + 1
    
    ready_to_assign = sum(1 for r in rows if r["ready_to_assign"] == "YES")
    urgent = sum(1 for r in rows if r["schedule_status"] == "URGENT")
    complete = sum(1 for r in rows if r["survey_complete"] == "YES")
    
    summary_ws['A1'] = "CEI SURVEY SUMMARY"
    summary_ws['A1'].font = Font(bold=True, size=14)
    
    summary_ws['A3'] = "Total CEI Sites"
    summary_ws['B3'] = total_sites
    
    summary_ws['A4'] = "Scout Submitted"
    summary_ws['B4'] = sum(1 for r in rows if r["scout_submitted"] == "YES")
    
    summary_ws['A5'] = "Survey Required"
    summary_ws['B5'] = surveys_required
    
    summary_ws['A6'] = "No Survey Needed"
    summary_ws['B6'] = no_survey_needed
    
    summary_ws['A7'] = "Pending Scout/Review"
    summary_ws['B7'] = pending
    
    summary_ws['A9'] = "By Survey Type:"
    summary_ws['A9'].font = Font(bold=True)
    
    row_idx = 10
    for survey_type in ["BOTH", "CCTV", "FA/INTRUSION", "NONE", "PENDING", "REVIEW"]:
        count = survey_type_counts.get(survey_type, 0)
        summary_ws[f'A{row_idx}'] = f"  {survey_type}"
        summary_ws[f'B{row_idx}'] = count
        if survey_type in survey_colors:
            summary_ws[f'B{row_idx}'].fill = PatternFill(
                start_color=survey_colors[survey_type],
                end_color=survey_colors[survey_type],
                fill_type="solid"
            )
        row_idx += 1
    
    summary_ws['A13'] = "Status Breakdown:"
    summary_ws['A13'].font = Font(bold=True)
    
    summary_ws['A14'] = "Ready to Assign"
    summary_ws['B14'] = ready_to_assign
    
    summary_ws['A15'] = "Urgent (≤165 days)"
    summary_ws['B15'] = urgent
    
    summary_ws['A16'] = "Already Complete"
    summary_ws['B16'] = complete
    
    # Adjust summary sheet widths
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 12
    
    # Save
    output_file = Path("CEI_Survey_Report.xlsx")
    wb.save(output_file)
    logger.info(f"✅ Report saved to {output_file}")
    
    # Print summary
    print("\n" + "="*70)
    print("CEI SURVEY REPORT SUMMARY")
    print("="*70)
    print(f"Total CEI Sites:        {total_sites}")
    print(f"Scout Submitted:        {sum(1 for r in rows if r['scout_submitted'] == 'YES')}")
    print(f"Survey Required:        {surveys_required}")
    print(f"No Survey Needed:       {no_survey_needed}")
    print(f"Pending Scout/Review:   {pending}")
    print(f"\nBy Survey Type:")
    for survey_type in ["BOTH", "CCTV", "FA/INTRUSION", "NONE", "PENDING", "REVIEW"]:
        count = survey_type_counts.get(survey_type, 0)
        if count > 0:
            print(f"  {survey_type:20s} {count:4d}")
    print(f"\nReady to Assign:        {ready_to_assign}")
    print(f"Urgent (≤165 days):     {urgent}")
    print(f"Already Complete:       {complete}")
    print("="*70)


if __name__ == "__main__":
    main()
