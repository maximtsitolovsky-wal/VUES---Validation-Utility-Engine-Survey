import openpyxl
from pathlib import Path

excel_path = Path(r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\Excel\Vendor ASSIGN. 4.2.26.xlsx')

# Sites to remove (they're BOTH/Full Survey, not Scout)
sites_to_remove = {9, 864, 1590, 2070, 2071, 2188, 2308, 3883}

print(f'Loading: {excel_path}')
wb = openpyxl.load_workbook(excel_path)
print(f'Current sheets: {wb.sheetnames}')

# Remove from CEI (sites 9, 864, 1590)
if 'CEI' in wb.sheetnames:
    ws = wb['CEI']
    rows_to_delete = []
    for row_idx in range(2, ws.max_row + 1):
        site_val = ws.cell(row=row_idx, column=1).value
        if site_val and int(site_val) in sites_to_remove:
            rows_to_delete.append(row_idx)
            print(f'CEI: Marking row {row_idx} (site {site_val}) for removal')
    
    # Delete rows in reverse order to maintain indices
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx)
        print(f'CEI: Deleted row {row_idx}')

# Delete PTSI sheet entirely (new sheet we created)
if 'PTSI' in wb.sheetnames:
    del wb['PTSI']
    print('Deleted PTSI sheet')

# Delete alarmhawaii sheet entirely (new sheet we created)
if 'alarmhawaii' in wb.sheetnames:
    del wb['alarmhawaii']
    print('Deleted alarmhawaii sheet')

print(f'\nSaving: {excel_path}')
wb.save(excel_path)
wb.close()

print('\nDone! Removed 8 BOTH/Full Survey sites from Scout assignments.')
print('Sheets now:', end=' ')

# Verify
wb = openpyxl.load_workbook(excel_path, read_only=True)
print(wb.sheetnames)
wb.close()
