#!/usr/bin/env python3
"""Check where 891 is missing."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

# Check Everon sheet in Vendor Assignment file
vendor_file = r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\Excel\Vendor ASSIGN. 4.2.26.xlsx'
print(f"Checking: {vendor_file}")
wb = openpyxl.load_workbook(vendor_file, read_only=True, data_only=True)
ws = wb['Everon']
everon_sites = []
for row in ws.iter_rows(min_row=2, max_col=1):
    if row[0].value:
        everon_sites.append(str(row[0].value).strip().lstrip('0'))
wb.close()
print(f"Everon assignments: {len(everon_sites)} sites")
print(f"891 in Everon list: {'891' in everon_sites}")

# Check Scout reference file  
scout_file = r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\ScoutSurveyLab.xlsm'
print(f"\nChecking: {scout_file}")
wb2 = openpyxl.load_workbook(scout_file, read_only=True, data_only=True)
ws2 = wb2['Scout Map Data']
scout_sites = []
for row in ws2.iter_rows(min_row=2, max_col=1):
    if row[0].value:
        scout_sites.append(str(row[0].value).strip().lstrip('0'))
wb2.close()
print(f"Scout Map Data: {len(scout_sites)} sites")
print(f"891 in Scout Map Data: {'891' in scout_sites}")

if '891' in everon_sites and '891' not in scout_sites:
    print("\n⚠️  PROBLEM: Site 891 is in Everon assignments but MISSING from Scout Map Data!")
    print("   → Need to add 891 to ScoutSurveyLab.xlsm 'Scout Map Data' sheet")
