#!/usr/bin/env python
"""Compare sites between the two Excel files to understand the mapping."""
import openpyxl
import sys
sys.path.insert(0, 'src')

# Load vendor assignment file (5 vendors)
assign_path = r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\Excel\Vendor ASSIGN. 4.2.26.xlsx'
wb1 = openpyxl.load_workbook(assign_path, data_only=True)

assign_sites = {}  # site -> vendor
for sheet in wb1.sheetnames:
    ws = wb1[sheet]
    vendor = sheet
    for row in range(2, ws.max_row + 1):
        site = ws.cell(row, 1).value
        if site:
            site = str(site).strip().lstrip('0')
            if site.isdigit():
                assign_sites[site] = vendor

print(f'Vendor Assignment File: {len(assign_sites)} unique sites')
for v in ['Wachter', 'Techwise', 'SAS', 'Everon', 'CEI']:
    count = sum(1 for s, vnd in assign_sites.items() if vnd == v)
    print(f'  {v}: {count}')

# Load survey routing file (3 vendors)
from siteowlqa.survey_routing import load_schedule_data, DEFAULT_WORKBOOK_PATH
schedule_records = load_schedule_data(DEFAULT_WORKBOOK_PATH)

schedule_sites = {}  # site -> vendor
for r in schedule_records:
    if r.site:
        schedule_sites[r.site] = r.vendor or 'EMPTY'

print(f'\nSurvey Lab File: {len(schedule_sites)} unique sites')
for v in ['Wachter', 'Everon', 'CEI', 'EMPTY']:
    count = sum(1 for s, vnd in schedule_sites.items() if vnd == v)
    print(f'  {v}: {count}')

# Cross-reference: Where did Techwise/SAS sites go in Survey Lab?
techwise_sites = [s for s, v in assign_sites.items() if v == 'Techwise']
sas_sites = [s for s, v in assign_sites.items() if v == 'SAS']

print(f'\n=== Cross-Reference ===')
print(f'Techwise sites in assignment file: {len(techwise_sites)}')

techwise_in_lab = {}
for site in techwise_sites:
    if site in schedule_sites:
        v = schedule_sites[site]
        techwise_in_lab[v] = techwise_in_lab.get(v, 0) + 1

print(f'Techwise sites mapped to in Survey Lab:')
for v, count in techwise_in_lab.items():
    print(f'  -> {v}: {count}')

techwise_missing = [s for s in techwise_sites if s not in schedule_sites]
print(f'Techwise sites NOT in Survey Lab: {len(techwise_missing)}')
if techwise_missing[:5]:
    print(f'  Examples: {techwise_missing[:5]}')

print(f'\nSAS sites in assignment file: {len(sas_sites)}')
sas_in_lab = {}
for site in sas_sites:
    if site in schedule_sites:
        v = schedule_sites[site]
        sas_in_lab[v] = sas_in_lab.get(v, 0) + 1

print(f'SAS sites mapped to in Survey Lab:')
for v, count in sas_in_lab.items():
    print(f'  -> {v}: {count}')

sas_missing = [s for s in sas_sites if s not in schedule_sites]
print(f'SAS sites NOT in Survey Lab: {len(sas_missing)}')
if sas_missing[:5]:
    print(f'  Examples: {sas_missing[:5]}')

# Save
with open('_site_comparison.txt', 'w') as f:
    f.write('=== SITE COMPARISON ===\n\n')
    f.write(f'Vendor Assignment File: {len(assign_sites)} sites\n')
    f.write(f'Survey Lab File: {len(schedule_sites)} sites\n\n')
    f.write(f'Techwise ({len(techwise_sites)} sites) mapped to:\n')
    for v, count in techwise_in_lab.items():
        f.write(f'  {v}: {count}\n')
    f.write(f'  Missing: {len(techwise_missing)}\n\n')
    f.write(f'SAS ({len(sas_sites)} sites) mapped to:\n')
    for v, count in sas_in_lab.items():
        f.write(f'  {v}: {count}\n')
    f.write(f'  Missing: {len(sas_missing)}\n')

print('\nSaved to _site_comparison.txt')
