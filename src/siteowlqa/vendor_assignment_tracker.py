"""
Vendor Assignment Tracker
==========================
Reads vendor assignments from Excel and compares with Scout Airtable completions
to track remaining assignments and completion velocity.

Features:
- Load vendor assignments from Excel workbook
- Compare against Scout Airtable completed submissions
- Calculate remaining assignments per vendor
- Track completion velocity (time / total assignments)
- Provide vendor performance insights for dashboard and weekly highlights
"""
import logging
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class VendorAssignment:
    """Represents a vendor assignment from the Excel file."""
    site_number: str
    vendor_name: str
    assigned_date: Optional[datetime] = None
    
    
@dataclass
class VendorStats:
    """Statistics for a single vendor."""
    vendor_name: str
    total_assigned: int
    completed: int
    remaining: int
    completion_rate: float  # percentage
    avg_completion_days: Optional[float] = None
    
    def to_dict(self) -> dict:
        """Convert to dictionary for JSON serialization."""
        return {
            "vendor_name": self.vendor_name,
            "total_assigned": self.total_assigned,
            "completed": self.completed,
            "remaining": self.remaining,
            "completion_rate": round(self.completion_rate, 2),
            "avg_completion_days": round(self.avg_completion_days, 2) if self.avg_completion_days else None
        }


class VendorAssignmentTracker:
    """Tracks vendor assignments and completion progress."""
    
    # Known vendor names (normalized)
    KNOWN_VENDORS = ["Wachter", "Techwise", "SAS", "Everon", "CEI"]
    
    def __init__(self, assignment_file_path: str):
        """
        Initialize tracker with path to vendor assignment Excel file.
        
        Args:
            assignment_file_path: Path to the vendor assignment Excel file
        """
        self.assignment_file = Path(assignment_file_path)
        self.assignments: List[VendorAssignment] = []
        self._loaded = False
    
    def load_assignments(self, sheet_name: str = "Sheet1") -> bool:
        """
        Load vendor assignments from Excel file.
        NOW LOADS ALL SHEETS to capture all vendors!
        
        Args:
            sheet_name: Ignored - we load all sheets
            
        Returns:
            True if loaded successfully, False otherwise
        """
        if not self.assignment_file.exists():
            logger.error(f"Assignment file not found: {self.assignment_file}")
            return False
        
        try:
            logger.info(f"Loading vendor assignments from {self.assignment_file.name}...")
            wb = openpyxl.load_workbook(self.assignment_file, data_only=True)
            
            total_assignments_loaded = 0
            
            # Load ALL sheets in the workbook - each sheet = one vendor
            for sheet_name in wb.sheetnames:
                logger.info(f"Processing sheet: {sheet_name}")
                ws = wb[sheet_name]
                
                # Skip if sheet is empty
                if ws.max_row < 2:
                    logger.info(f"Skipping empty sheet: {sheet_name}")
                    continue
                
                # Use sheet name as the vendor name
                vendor_name = self._normalize_vendor_name(sheet_name)
                logger.info(f"Using vendor name from sheet: {vendor_name}")
            
                # Find header row (look for "Site Number" or "Store Number")
                header_row = None
                site_col_idx = None
                date_col_idx = None
                
                for row_idx in range(1, min(21, ws.max_row + 1)):
                    for col_idx in range(1, min(21, ws.max_column + 1)):
                        cell_value = ws.cell(row_idx, col_idx).value
                        if cell_value:
                            cell_str = str(cell_value).strip().lower()
                            
                            if "site" in cell_str or "store" in cell_str:
                                if "number" in cell_str:
                                    header_row = row_idx
                                    site_col_idx = col_idx
                            elif "date" in cell_str or "assigned" in cell_str:
                                date_col_idx = col_idx
                    
                    if header_row and site_col_idx:
                        break
                
                if not (header_row and site_col_idx):
                    logger.warning(f"Sheet '{sheet_name}': Could not find Site Number column, skipping")
                    continue
                
                logger.info(f"Sheet '{sheet_name}': Found headers at row {header_row}, Site col={site_col_idx}")
                
                # Load assignments from this sheet
                sheet_assignments_loaded = 0
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    site_num = ws.cell(row_idx, site_col_idx).value
                    assigned_date = ws.cell(row_idx, date_col_idx).value if date_col_idx else None
                    
                    if site_num:
                        site_num_str = self._normalize_site_number(str(site_num).strip())
                        
                        # Parse date if available
                        date_obj = None
                        if assigned_date:
                            if isinstance(assigned_date, datetime):
                                date_obj = assigned_date
                            elif isinstance(assigned_date, str):
                                try:
                                    date_obj = datetime.strptime(assigned_date, "%Y-%m-%d")
                                except ValueError:
                                    try:
                                        date_obj = datetime.strptime(assigned_date, "%m/%d/%Y")
                                    except ValueError:
                                        pass
                        
                        self.assignments.append(VendorAssignment(
                            site_number=site_num_str,
                            vendor_name=vendor_name,  # Use sheet name as vendor
                            assigned_date=date_obj
                        ))
                        sheet_assignments_loaded += 1
                
                logger.info(f"Sheet '{sheet_name}': Loaded {sheet_assignments_loaded} assignments for vendor '{vendor_name}'")
                total_assignments_loaded += sheet_assignments_loaded
            
            wb.close()
            logger.info(f"Loaded {total_assignments_loaded} total vendor assignments from {len(wb.sheetnames)} sheets")
            self._loaded = True
            return True
            
        except Exception as e:
            logger.error(f"Error loading vendor assignments: {e}")
            return False
    
    def _normalize_vendor_name(self, vendor: str) -> str:
        """Normalize vendor name to match known vendors."""
        vendor_upper = vendor.upper()
        
        for known in self.KNOWN_VENDORS:
            if known.upper() in vendor_upper:
                return known
        
        return vendor  # Return as-is if not recognized
    
    def _normalize_site_number(self, site_num: str) -> str:
        """Normalize site number by removing leading zeros for consistent matching."""
        try:
            # Convert to int and back to str to remove leading zeros
            return str(int(site_num))
        except (ValueError, TypeError):
            # If conversion fails, return as-is
            return str(site_num).strip()

    def _record_sort_key(self, record: dict[str, Any]) -> tuple[int, str, str, str]:
        """Prefer completed Scout responses, then newest dates."""
        raw_fields = record.get("raw_fields", {}) or {}
        is_complete = 1 if raw_fields.get("Complete?") == 1 else 0
        scout_date = str(raw_fields.get("Scout Date") or record.get("submitted_at") or "")
        created_time = str(record.get("created_time") or "")
        record_id = str(record.get("record_id") or "")
        return (is_complete, scout_date, created_time, record_id)

    def _pick_best_records_by_site(
        self,
        scout_records: List[dict[str, Any]],
    ) -> dict[str, dict[str, Any]]:
        """Collapse multiple Airtable records per site into one best live record."""
        records_by_site: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for record in scout_records:
            site_number = self._normalize_site_number(str(record.get("site_number", "")).strip())
            if site_number:
                records_by_site[site_number].append(record)

        return {
            site_number: max(records, key=self._record_sort_key)
            for site_number, records in records_by_site.items()
        }

    def _build_response_summary(self, raw_fields: dict[str, Any]) -> list[str]:
        """Return a compact Scout-response summary for the live mesh table."""
        field_map = [
            ("Panel", "Fire Panel Type"),
            ("Tri-Mount", "Rooftop Tri-Mount Present"),
            ("Baluns", "Analog CCTV Baluns Present?"),
            ("Cable", "Coax or Siamese Cable"),
            ("Homerun", "Homerun Cabling Present"),
            ("Condition", "Cable Condition"),
        ]
        summary: list[str] = []
        for label, field_name in field_map:
            value = raw_fields.get(field_name)
            if value in (None, "", []):
                continue
            if isinstance(value, list):
                value = ", ".join(str(item) for item in value if item not in (None, ""))
            summary.append(f"{label}: {value}")
        return summary[:4]

    def _build_mesh_row(
        self,
        *,
        assignment: VendorAssignment | None,
        record: dict[str, Any] | None,
    ) -> dict[str, Any]:
        """Build one row of Excel × Airtable mesh data for the UI."""
        raw_fields = record.get("raw_fields", {}) if record else {}
        assigned_vendor = assignment.vendor_name if assignment else ""
        assigned_date = assignment.assigned_date.isoformat() if assignment and assignment.assigned_date else ""
        site_number = assignment.site_number if assignment else self._normalize_site_number(
            str(record.get("site_number", "")).strip()
        )

        completed_by_vendor = self._normalize_vendor_name(
            str(raw_fields.get("Surveyor Parent Company", "")).strip()
        ) if record else ""
        is_complete = raw_fields.get("Complete?") == 1 if record else False

        status_key = "pending_assignment"
        status_label = "Pending in Excel"
        status_rank = 2

        if assignment and record:
            if is_complete and completed_by_vendor == assigned_vendor:
                status_key = "matched_complete"
                status_label = "Win · matched complete"
                status_rank = 0
            elif is_complete and completed_by_vendor and completed_by_vendor != assigned_vendor:
                status_key = "completed_by_other_vendor"
                status_label = "Loss · wrong vendor"
                status_rank = 1
            else:
                status_key = "started_not_complete"
                status_label = "Loss · response started"
                status_rank = 2
        elif assignment and not record:
            status_key = "pending_assignment"
            status_label = "Loss · no Airtable response"
            status_rank = 3
        elif record and not assignment:
            status_key = "unassigned_airtable"
            status_label = "Loss · Airtable only"
            status_rank = 4

        return {
            "site_number": site_number,
            "assigned_vendor": assigned_vendor or "—",
            "assigned_date": assigned_date,
            "record_id": str(record.get("record_id") or "") if record else "",
            "created_time": str(record.get("created_time") or "") if record else "",
            "status_key": status_key,
            "status_label": status_label,
            "status_rank": status_rank,
            "has_excel_assignment": assignment is not None,
            "has_airtable_record": record is not None,
            "is_complete": is_complete,
            "completed_by_vendor": completed_by_vendor or "—",
            "surveyor_name": str(raw_fields.get("Surveyor Name") or "").strip(),
            "surveyor_email": str(raw_fields.get("Surveyor Email") or "").strip(),
            "scout_date": str(raw_fields.get("Scout Date") or record.get("submitted_at") or "") if record else "",
            "comments": str(raw_fields.get("Comments") or "").strip(),
            "panel_type": str(raw_fields.get("Fire Panel Type") or "").strip(),
            "response_summary": self._build_response_summary(raw_fields),
        }

    def build_assignment_mesh(
        self,
        scout_records: List[dict[str, Any]],
    ) -> dict[str, Any]:
        """Join Excel assignments with live Airtable Scout responses for the UI."""
        if not self._loaded:
            return {
                "generated_at": datetime.utcnow().isoformat(),
                "rows": [],
                "summary": {
                    "matched_complete": 0,
                    "started_not_complete": 0,
                    "pending_assignment": 0,
                    "completed_by_other_vendor": 0,
                    "unassigned_airtable": 0,
                    "losses_total": 0,
                },
            }

        best_records = self._pick_best_records_by_site(scout_records)
        rows: list[dict[str, Any]] = []
        summary = {
            "matched_complete": 0,
            "started_not_complete": 0,
            "pending_assignment": 0,
            "completed_by_other_vendor": 0,
            "unassigned_airtable": 0,
            "losses_total": 0,
        }

        sorted_assignments = sorted(
            self.assignments,
            key=lambda item: (item.vendor_name, item.site_number),
        )
        for assignment in sorted_assignments:
            record = best_records.pop(assignment.site_number, None)
            row = self._build_mesh_row(assignment=assignment, record=record)
            summary[row["status_key"]] += 1
            rows.append(row)

        for site_number, record in sorted(best_records.items()):
            row = self._build_mesh_row(assignment=None, record=record)
            summary[row["status_key"]] += 1
            rows.append(row)

        summary["losses_total"] = (
            summary["started_not_complete"]
            + summary["pending_assignment"]
            + summary["completed_by_other_vendor"]
            + summary["unassigned_airtable"]
        )

        rows.sort(key=lambda row: (row["status_rank"], row["assigned_vendor"], row["site_number"]))
        return {
            "generated_at": datetime.utcnow().isoformat(),
            "rows": rows,
            "summary": summary,
        }
    
    def calculate_vendor_stats(
        self, 
        completed_submissions: List[dict]
    ) -> Dict[str, VendorStats]:
        """
        Calculate statistics for each vendor based on assignments and completions.
        
        Args:
            completed_submissions: List of completed Scout submissions from Airtable.
                Each dict should have: site_number, vendor_name, submitted_at
        
        Returns:
            Dictionary mapping vendor name to VendorStats
        """
        if not self._loaded:
            logger.warning("Assignments not loaded yet")
            return {}
        
        # Build completion map by site number (normalized)
        completion_map = {}
        for sub in completed_submissions:
            site_num = self._normalize_site_number(str(sub.get("site_number", "")).strip())
            vendor = self._normalize_vendor_name(str(sub.get("vendor_name", "")).strip())
            submitted_at = sub.get("submitted_at")
            
            if site_num:
                completion_map[site_num] = {
                    "vendor": vendor,
                    "submitted_at": submitted_at
                }
        
        # Group assignments by vendor
        vendor_assignments: Dict[str, List[VendorAssignment]] = {}
        for assignment in self.assignments:
            vendor = assignment.vendor_name
            if vendor not in vendor_assignments:
                vendor_assignments[vendor] = []
            vendor_assignments[vendor].append(assignment)
        
        # Calculate stats for each vendor
        vendor_stats = {}
        
        for vendor, assignments in vendor_assignments.items():
            total_assigned = len(assignments)
            completed_count = 0
            completion_days = []
            
            for assignment in assignments:
                # Only count as completed if the SAME vendor completed it
                completion_info = completion_map.get(assignment.site_number)
                if completion_info:
                    completing_vendor = self._normalize_vendor_name(completion_info.get("vendor", ""))
                    # Must be completed by the assigned vendor
                    if completing_vendor == vendor:
                        completed_count += 1
                        
                        # Calculate completion time if we have dates
                        if assignment.assigned_date:
                            submitted_at = completion_info.get("submitted_at")
                            
                            if submitted_at:
                                if isinstance(submitted_at, str):
                                    try:
                                        submitted_at = datetime.fromisoformat(submitted_at.replace("Z", "+00:00"))
                                    except:
                                        try:
                                            submitted_at = datetime.strptime(submitted_at, "%Y-%m-%d")
                                        except:
                                            submitted_at = None
                                
                                if submitted_at:
                                    days = (submitted_at - assignment.assigned_date).days
                                    if days >= 0:  # Only count positive days
                                        completion_days.append(days)
            
            remaining = total_assigned - completed_count
            completion_rate = (completed_count / total_assigned * 100) if total_assigned > 0 else 0.0
            avg_days = sum(completion_days) / len(completion_days) if completion_days else None
            
            vendor_stats[vendor] = VendorStats(
                vendor_name=vendor,
                total_assigned=total_assigned,
                completed=completed_count,
                remaining=remaining,
                completion_rate=completion_rate,
                avg_completion_days=avg_days
            )
        
        return vendor_stats
    
    def get_summary(self) -> dict:
        """Get summary of all assignments."""
        if not self._loaded:
            return {
                "total_assignments": 0,
                "vendors": [],
                "loaded": False
            }
        
        vendor_counts = {}
        for assignment in self.assignments:
            vendor = assignment.vendor_name
            vendor_counts[vendor] = vendor_counts.get(vendor, 0) + 1
        
        return {
            "total_assignments": len(self.assignments),
            "vendors": [
                {"name": vendor, "assignments": count}
                for vendor, count in sorted(vendor_counts.items(), key=lambda x: -x[1])
            ],
            "loaded": True
        }
