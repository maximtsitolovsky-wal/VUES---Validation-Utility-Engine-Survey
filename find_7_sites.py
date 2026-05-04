import json
import openpyxl
from pathlib import Path

# The actual Scout reference file path
scout_ref_path = Path(r"C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\ScoutSurveyLab.xlsm")

print(f'Loading Scout Map Data from: {scout_ref_path}')
print(f'Exists: {scout_ref_path.exists()}')

scout_map_sites = set()

if scout_ref_path.exists():
    wb = openpyxl.load_workbook(scout_ref_path, read_only=True, data_only=True)
    print(f'Sheets: {wb.sheetnames}')
    
    if "Scout Map Data" in wb.sheetnames:
        ws = wb["Scout Map Data"]
        for row in ws.iter_rows(min_row=2, max_col=1):
            if row[0].value:
                site = str(row[0].value).strip().lstrip("0")
                scout_map_sites.add(site)
        print(f'Sites in Scout Map Data: {len(scout_map_sites)}')
    else:
        print('Sheet "Scout Map Data" not found')
    wb.close()

# Load vendor assignments
with open('output/team_dashboard_data.json') as f:
    data = json.load(f)

va = data.get('vendor_assignments', {})
mesh_data = va.get('mesh', {})
mesh_rows = mesh_data.get('rows', []) if isinstance(mesh_data, dict) else []

assigned_sites = set()
for row in mesh_rows:
    site = str(row.get('site_number', '')).strip().lstrip("0")
    if site:
        assigned_sites.add(site)

print(f'Sites in Vendor Assignments: {len(assigned_sites)}')

# Find the gap
if scout_map_sites:
    unassigned = scout_map_sites - assigned_sites
    print(f'\n=== {len(unassigned)} SITES IN SCOUT MAP BUT NOT IN VENDOR ASSIGN ===')
    for site in sorted(unassigned, key=lambda x: int(x) if x.isdigit() else 0):
        print(f'  Site {site}')
    
    # Also check reverse
    extra_assigned = assigned_sites - scout_map_sites
    if extra_assigned:
        print(f'\n=== {len(extra_assigned)} SITES ASSIGNED BUT NOT IN SCOUT MAP ===')
        for site in sorted(extra_assigned, key=lambda x: int(x) if x.isdigit() else 0):
            print(f'  Site {site}')
