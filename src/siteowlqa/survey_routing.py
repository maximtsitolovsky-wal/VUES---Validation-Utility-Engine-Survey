"""survey_routing.py — Survey Routing + Vendor Output Sheet logic.

Evaluates scout responses to determine:
- Whether each site needs a survey
- Survey type: CCTV, FA/INTRUSION, BOTH, NONE, or REVIEW
- Full upgrade scenarios
- Supplemental flags
- 165-day construction scheduling

Data Sources:
- Airtable Scout Table (source of truth for scout answers)
- Excel workbook (vendor ownership and timing)
"""

from __future__ import annotations

import json
import logging
import re
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Any

import requests

from siteowlqa.config import (
    SURVEY_TYPE_CCTV,
    SURVEY_TYPE_FA_INTRUSION,
    SURVEY_TYPE_BOTH,
)

log = logging.getLogger(__name__)

# Airtable Scout Table
SCOUT_BASE_ID = "appAwgaX89x0JxG3Z"
SCOUT_TABLE_ID = "tblC4o9AvVulyxFMk"

# Airtable Survey Routing Table (same base, different table)
ROUTING_TABLE_ID = "tbl4LbgPUluSrbG2K"

# Airtable Survey Submissions Table (different base - for tracking completed surveys)
SURVEY_SUBMISSIONS_BASE_ID = "apptK6zNN0Hf3OuoJ"
SURVEY_SUBMISSIONS_TABLE_ID = "tblo5JLmY0XhigcMO"

# Excel workbook path
DEFAULT_WORKBOOK_PATH = (
    r"C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\2027 Survey Lab.xlsm"
)

# Scheduling threshold
CONSTRUCTION_DEADLINE_DAYS = 165

# Valid survey vendors (only these should appear in output)
VALID_SURVEY_VENDORS = {"Wachter", "CEI", "Everon"}

# Vendor reassignment mapping: these vendors' sites are surveyed by CEI
# (Techwise and SAS assignments transferred to CEI per 2026-05-04 decision)
VENDOR_REASSIGNMENT = {
    "techwise": "CEI",
    "sas": "CEI",
}

# Invalid values to filter out
INVALID_VALUES = {"#REF!", "#N/A", "#VALUE!", "#ERROR!", "MAXIM"}

# Manual scout overrides: sites where scout is complete and needs BOTH surveys
# Format: site -> survey_type
MANUAL_SCOUT_COMPLETE = {
    "864": "BOTH",
    "9": "BOTH",
}


@dataclass
class SurveyRoutingRow:
    """Final output row for vendor-ready table."""
    site: str
    vendor: str
    days_to_construction: str
    survey_required: str  # YES, NO, PENDING, REVIEW
    survey_type: str  # CCTV, FA/INTRUSION, BOTH, NONE, PENDING, REVIEW
    upgrade_decision: str
    reason_for_decision: str
    schedule_status: str  # NOT REQUIRED, ON TRACK, URGENT, COMPLETE, PENDING, REVIEW
    ready_to_assign: str  # YES, NO
    supplemental_flags: str
    vendor_instructions: str
    survey_complete: bool = False
    # Progress tracking from Excel
    assigned: bool = False
    assigned_vendor: str = ""
    survey_returned_qa: bool = False
    passes_at_qa: bool = False
    survey_returned_date: str = ""
    in_siteowl_design: bool = False
    in_siteowl_installation: bool = False
    in_siteowl_live: bool = False
    percentage_complete: float = 0.0
    on_project_tracking: bool = True
    # Tracking fields
    scout_submitted: bool = False
    notes: str = ""


@dataclass
class ScoutAnswers:
    """Parsed scout answers from Airtable."""
    site: str
    record_id: str
    # FA/Intrusion fields
    one_notification_device: bool = False
    ceiling_mounted_devices: bool = False
    sales_floor_column_devices: bool = False
    emergency_exit_only_devices: bool = False
    fire_panel_type: str = ""
    # CCTV fields
    coax_siamese_cable: bool = False
    analog_baluns_present: bool = False
    rooftop_trimount_present: bool = False
    rooftop_trimount_count: int = 0
    cable_condition: str = ""
    homerun_cabling_present: bool = False
    ap_office_moving: str = ""  # YES, NO, or blank
    raw_fields: dict = field(default_factory=dict)


@dataclass
class ScheduleData:
    """Data from Excel workbook - Project Tracking + MAP DATA."""
    site: str
    vendor: str  # From MAP DATA Vendor_Final
    days_to_construction: int | None
    # Progress tracking from Project Tracking
    assigned: bool = False  # Col 6
    assigned_vendor: str = ""  # Col 8
    survey_returned_qa: bool = False  # Col 12
    passes_at_qa: bool = False  # Col 13
    survey_returned_date: str = ""  # Col 14
    in_siteowl_design: bool = False  # Col 16
    in_siteowl_installation: bool = False  # Col 17
    in_siteowl_live: bool = False  # Col 21
    percentage_complete: float = 0.0  # Col 22
    survey_complete: bool = False  # Derived from progress columns
    on_project_tracking: bool = True  # Is this site on the Project Tracking sheet?

def _normalize_site(site: str | int | float | None) -> str:
    """Normalize site ID for consistent joining.
    
    Strips leading zeros so '0336' matches '336'.
    """
    if site is None:
        return ""
    site_str = str(site).strip()
    # Remove .0 from float conversion
    if site_str.endswith(".0"):
        site_str = site_str[:-2]
    # Strip leading zeros for consistent matching
    # But keep at least one digit (so "0" stays "0")
    site_str = site_str.lstrip("0") or "0"
    return site_str


def _is_yes(value: Any) -> bool:
    """Check if a value represents YES/True."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().upper() in ("YES", "TRUE", "1", "Y")
    return bool(value)


def _is_blank(value: Any) -> bool:
    """Check if a value is blank/empty."""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def fetch_scout_data(token: str) -> list[ScoutAnswers]:
    """Fetch scout responses from Airtable."""
    url = f"https://api.airtable.com/v0/{SCOUT_BASE_ID}/{SCOUT_TABLE_ID}"
    headers = {"Authorization": f"Bearer {token}"}
    
    all_records = []
    offset = None
    
    while True:
        params = {"pageSize": 100}
        if offset:
            params["offset"] = offset
            
        try:
            resp = requests.get(url, headers=headers, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            log.error(f"Failed to fetch scout data from Airtable: {e}")
            break
            
        records = data.get("records", [])
        all_records.extend(records)
        
        offset = data.get("offset")
        if not offset:
            break
    
    log.info(f"Fetched {len(all_records)} scout records from Airtable")
    
    # Parse into ScoutAnswers objects
    parsed = []
    for rec in all_records:
        fields = rec.get("fields", {})
        site = _normalize_site(fields.get("Site Number") or fields.get("Site") or fields.get("site"))
        
        if not site:
            continue
            
        # Map Airtable field names to our structure
        # These field names need to match the actual Airtable schema EXACTLY
        answers = ScoutAnswers(
            site=site,
            record_id=rec.get("id", ""),
            # FA/Intrusion
            one_notification_device=_is_yes(fields.get("Does this site only have one notification device located in the store? (Usually at the Service Desk)")),
            ceiling_mounted_devices=_is_yes(fields.get("Are any notification devices ceiling mounted?")),
            sales_floor_column_devices=_is_yes(fields.get("Are any notification devices mounted on sales floor columns?")),
            emergency_exit_only_devices=_is_yes(fields.get("Are notification devices installed only above emergency exits?")),
            fire_panel_type=str(fields.get("Fire Panel Type", "") or ""),
            # CCTV
            coax_siamese_cable=_is_yes(fields.get("Coax or Siamese Cable")),
            analog_baluns_present=_is_yes(fields.get("Analog CCTV Baluns Present?")),
            rooftop_trimount_present=_is_yes(fields.get("Rooftop Tri-Mount Present")),
            rooftop_trimount_count=int(fields.get("Rooftop Tri-Mount Count") or 0) if fields.get("Rooftop Tri-Mount Count") else 0,
            cable_condition=str(fields.get("Cable Condition", "") or ""),
            homerun_cabling_present=_is_yes(fields.get("Homerun Cabling Present")),
            ap_office_moving=str(fields.get("AP Office Moving", "") or "").strip().upper(),
            raw_fields=fields,
        )
        parsed.append(answers)
    
    return parsed


def fetch_survey_submissions(token: str) -> set[str]:
    """Fetch completed survey submissions from Airtable Submissions table.
    
    Returns set of site numbers that have PASS status surveys.
    """
    url = f"https://api.airtable.com/v0/{SURVEY_SUBMISSIONS_BASE_ID}/{SURVEY_SUBMISSIONS_TABLE_ID}"
    headers = {"Authorization": f"Bearer {token}"}
    
    completed_sites = set()
    offset = None
    
    while True:
        params = {
            "pageSize": 100,
            # Only fetch PASS records to reduce data transfer
            "filterByFormula": "{Processing Status} = 'PASS'",
        }
        if offset:
            params["offset"] = offset
            
        try:
            resp = requests.get(url, headers=headers, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            log.error(f"Failed to fetch survey submissions from Airtable: {e}")
            break
            
        records = data.get("records", [])
        for rec in records:
            fields = rec.get("fields", {})
            # Try multiple field names for site number
            site = _normalize_site(
                fields.get("Site Number") or 
                fields.get("Site") or 
                fields.get("site") or
                fields.get("Site #") or
                fields.get("Store Number")
            )
            if site:
                completed_sites.add(site)
        
        offset = data.get("offset")
        if not offset:
            break
    
    log.info(f"Fetched {len(completed_sites)} completed survey sites from Airtable Submissions")
    return completed_sites


def _normalize_vendor(vendor: str | None) -> str:
    """Normalize vendor name - case insensitive, trim whitespace.
    
    Also applies vendor reassignment (Techwise/SAS -> CEI).
    """
    if not vendor:
        return ""
    v = str(vendor).strip()
    if v in INVALID_VALUES or v.startswith("#"):
        return ""
    v_lower = v.lower()
    
    # Check if vendor should be reassigned (e.g., Techwise/SAS -> CEI)
    if v_lower in VENDOR_REASSIGNMENT:
        return VENDOR_REASSIGNMENT[v_lower]
    
    # Case-insensitive match against valid vendors
    for valid in VALID_SURVEY_VENDORS:
        if v_lower == valid.lower():
            return valid  # Return canonical casing
    return ""  # Not a valid survey vendor


def load_map_data(workbook_path: str | Path) -> dict[str, str]:
    """Load vendor assignments from MAP DATA sheet.
    
    Returns dict mapping site -> vendor.
    """
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl not installed")
        return {}
    
    path = Path(workbook_path)
    if not path.exists():
        return {}
    
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        log.error(f"Failed to open workbook {path}: {e}")
        return {}
    
    if "MAP DATA" not in wb.sheetnames:
        log.warning(f"Sheet 'MAP DATA' not found in {path}")
        wb.close()
        return {}
    
    ws = wb["MAP DATA"]
    vendor_map = {}
    skipped_short = 0
    
    # Col 0 = Store Number, Col 12 = Vendor_Final, Col 15 = Vendor_Assigned
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        
        # RO-002 FIX: Allow short rows, just check what we need
        site = _normalize_site(row[0]) if len(row) > 0 else ""
        if not site:
            continue
        
        # Prefer Vendor_Final (col 12), fallback to Vendor_Assigned (col 15)
        vendor = ""
        if len(row) > 12:
            vendor = _normalize_vendor(row[12])
        if not vendor and len(row) > 15:
            vendor = _normalize_vendor(row[15])
        
        if vendor:
            vendor_map[site] = vendor
    
    wb.close()
    log.info(f"Loaded {len(vendor_map)} vendor assignments from MAP DATA")
    return vendor_map


def load_schedule_data(workbook_path: str | Path) -> list[ScheduleData]:
    """Load progress tracking data from Project Tracking sheet.
    
    Columns:
    - 0: StoreId
    - 6: Assigned (Y/N)
    - 8: Assigned Vendor (QB)
    - 12: Survey Returned for QA (AT) (Y/N)
    - 13: Passes AT QA (Y/N)
    - 14: Site Survey Returned Date (QB)
    - 16: Survey in Siteowl (Design)?(Y/N)
    - 17: Survey in Siteowl (Installation)?(Y/N)
    - 21: Survey in SiteOwl (Live Sites)
    - 22: Percentage to Complete
    - 23: Days to Construction Count Down
    - 24: Map Data Vendor Survey
    """
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl not installed - cannot read Excel workbook")
        return []
    
    path = Path(workbook_path)
    if not path.exists():
        log.warning(f"Workbook not found: {path}")
        return []
    
    # Load MAP DATA vendors first
    vendor_map = load_map_data(workbook_path)
    
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        log.error(f"Failed to open workbook {path}: {e}")
        return []
    
    if "Project Tracking" not in wb.sheetnames:
        log.warning(f"Sheet 'Project Tracking' not found in {path}")
        wb.close()
        return []
    
    ws = wb["Project Tracking"]
    data = []
    skipped = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        if not row:
            continue
        
        # RO-001 FIX: Don't skip short rows - openpyxl read_only mode
        # doesn't append trailing empty cells. Check each column individually.
        site = _normalize_site(row[0]) if len(row) > 0 else ""  # Column A = StoreId
        if not site:
            continue
        
        # Skip invalid site values
        if site in INVALID_VALUES or site.startswith("#"):
            skipped += 1
            continue
        
        # Get vendor from MAP DATA (preferred) or fallback to col 24
        vendor = vendor_map.get(site, "")
        if not vendor and len(row) > 24:
            vendor = _normalize_vendor(row[24])
        
        # Parse all tracking columns
        assigned = _is_yes(row[6]) if len(row) > 6 else False
        assigned_vendor = str(row[8] or "").strip() if len(row) > 8 else ""
        survey_returned_qa = _is_yes(row[12]) if len(row) > 12 else False
        passes_at_qa = _is_yes(row[13]) if len(row) > 13 else False
        survey_returned_date = str(row[14] or "")[:10] if len(row) > 14 else ""
        in_siteowl_design = _is_yes(row[16]) if len(row) > 16 else False
        in_siteowl_installation = _is_yes(row[17]) if len(row) > 17 else False
        in_siteowl_live = _is_yes(row[21]) if len(row) > 21 else False
        
        # Percentage complete
        pct = 0.0
        if len(row) > 22 and row[22] is not None:
            try:
                pct = float(row[22])
                if pct > 1:  # Handle 0-100 vs 0-1 format
                    pct = pct / 100
            except (ValueError, TypeError):
                pass
        
        # Days to construction
        days = None
        if len(row) > 23 and row[23] is not None:
            try:
                days = int(float(row[23]))
            except (ValueError, TypeError):
                pass
        
        # Survey is complete if it passes AT QA or is in SiteOwl Live
        survey_complete = passes_at_qa or in_siteowl_live
        
        data.append(ScheduleData(
            site=site,
            vendor=vendor,
            days_to_construction=days,
            assigned=assigned,
            assigned_vendor=assigned_vendor,
            survey_returned_qa=survey_returned_qa,
            passes_at_qa=passes_at_qa,
            survey_returned_date=survey_returned_date,
            in_siteowl_design=in_siteowl_design,
            in_siteowl_installation=in_siteowl_installation,
            in_siteowl_live=in_siteowl_live,
            percentage_complete=pct,
            survey_complete=survey_complete,
            on_project_tracking=True,
        ))
    
    wb.close()
    if skipped > 0:
        log.info(f"Skipped {skipped} invalid rows from Excel")
    log.info(f"Loaded {len(data)} schedule rows from Project Tracking")
    return data


def evaluate_site(scout: ScoutAnswers | None, schedule: ScheduleData | None) -> SurveyRoutingRow:
    """Apply all routing logic to determine survey requirements.
    
    Logic order:
    1. Full upgrade triggers (override survey need for that category)
    2. Survey triggers (for categories without full upgrade)
    3. Supplemental flags (informational)
    4. Scheduling logic
    5. Vendor instructions
    """
    
    # Determine site ID
    site = scout.site if scout else (schedule.site if schedule else "UNKNOWN")
    
    # Get schedule data
    vendor = schedule.vendor if schedule else ""
    days = schedule.days_to_construction if schedule else None
    days_str = str(days) if days is not None else ""
    survey_complete = schedule.survey_complete if schedule else False
    
    # Handle missing scout data - can't make a routing decision yet
    if scout is None:
        return SurveyRoutingRow(
            site=site,
            vendor=vendor,
            days_to_construction=days_str,
            survey_required="PENDING",
            survey_type="PENDING",
            upgrade_decision="AWAITING SCOUT",
            reason_for_decision="Scout not submitted. Cannot determine survey requirement.",
            schedule_status="PENDING",
            ready_to_assign="NO",
            supplemental_flags="",
            vendor_instructions="Scout submission required before survey routing decision.",
            survey_complete=survey_complete,
            # Progress tracking from schedule
            assigned=schedule.assigned if schedule else False,
            assigned_vendor=schedule.assigned_vendor if schedule else "",
            survey_returned_qa=schedule.survey_returned_qa if schedule else False,
            passes_at_qa=schedule.passes_at_qa if schedule else False,
            survey_returned_date=schedule.survey_returned_date if schedule else "",
            in_siteowl_design=schedule.in_siteowl_design if schedule else False,
            in_siteowl_installation=schedule.in_siteowl_installation if schedule else False,
            in_siteowl_live=schedule.in_siteowl_live if schedule else False,
            percentage_complete=schedule.percentage_complete if schedule else 0.0,
            on_project_tracking=schedule.on_project_tracking if schedule else False,
            scout_submitted=False,  # Scout NOT submitted
        )
    
    if schedule is None:
        vendor = ""
        days_str = ""
    
    supplemental_flags_list = []
    reasons = []
    
    # === STEP 1: Check Full Upgrade Triggers ===
    # These determine if a survey is NOT needed for that category
    
    # FA/Intrusion Full Upgrade: only one notification device = YES
    fa_full_upgrade = scout.one_notification_device
    
    # CCTV Full Upgrade: coax/siamese cable = YES
    cctv_full_upgrade = scout.coax_siamese_cable
    
    # Conditional CCTV Full Upgrade: homerun + AP office moving = YES
    if scout.homerun_cabling_present and not cctv_full_upgrade:
        if scout.ap_office_moving == "YES":
            cctv_full_upgrade = True
        elif _is_blank(scout.ap_office_moving):
            # AP status unknown - note it but don't block routing
            # Homerun cabling will trigger CCTV survey below
            supplemental_flags_list.append("AP office move status unknown — will be determined during survey.")
    
    # === STEP 2: Check Survey Triggers for categories WITHOUT full upgrade ===
    fa_survey_needed = False
    cctv_survey_needed = False
    
    # FA/Intrusion survey triggers (only check if no FA full upgrade)
    if not fa_full_upgrade:
        if scout.ceiling_mounted_devices:
            fa_survey_needed = True
            supplemental_flags_list.append("Ceiling mounted notification devices")
        if scout.sales_floor_column_devices:
            fa_survey_needed = True
            supplemental_flags_list.append("Notification devices on sales floor columns")
        if scout.emergency_exit_only_devices:
            fa_survey_needed = True
            supplemental_flags_list.append("Notification devices only above emergency exits")
        if scout.fire_panel_type:
            fa_survey_needed = True
            supplemental_flags_list.append(f"Fire panel type: {scout.fire_panel_type}")
    
    # CCTV survey triggers (only check if no CCTV full upgrade)
    if not cctv_full_upgrade:
        if scout.analog_baluns_present:
            cctv_survey_needed = True
            supplemental_flags_list.append("Analog CCTV baluns present")
        if scout.rooftop_trimount_present:
            cctv_survey_needed = True
            supplemental_flags_list.append("Rooftop tri-mount present")
        if scout.rooftop_trimount_count > 0:
            cctv_survey_needed = True
            supplemental_flags_list.append(f"Rooftop tri-mount count: {scout.rooftop_trimount_count}")
        if scout.cable_condition:
            cctv_survey_needed = True
            supplemental_flags_list.append(f"Cable condition: {scout.cable_condition}")
        if scout.homerun_cabling_present:
            # Homerun cabling triggers CCTV survey need (AP status determined during survey)
            cctv_survey_needed = True
            supplemental_flags_list.append("Homerun cabling present")
    
    # === STEP 3: Determine Final Survey Type and Decision ===
    
    # Both full upgrades - no survey at all
    if fa_full_upgrade and cctv_full_upgrade:
        survey_required = "NO"
        survey_type = "NONE"
        upgrade_decision = "FULL BOTH UPGRADE"
        reason = "Both FA/Intrusion and CCTV full upgrade triggers present. No survey needed."
    
    # FA full upgrade only - check if CCTV survey still needed
    elif fa_full_upgrade and not cctv_full_upgrade:
        if cctv_survey_needed:
            survey_required = "YES"
            survey_type = SURVEY_TYPE_CCTV
            upgrade_decision = "FULL FA/INTRUSION UPGRADE + CCTV SURVEY"
            reason = "FA/Intrusion full upgrade (one notification device). CCTV survey still required."
        else:
            survey_required = "NO"
            survey_type = "NONE"
            upgrade_decision = "FULL FA/INTRUSION UPGRADE"
            reason = "Site has only one notification device. Full FA/Intrusion upgrade required. No survey needed."
    
    # CCTV full upgrade only - check if FA survey still needed
    elif cctv_full_upgrade and not fa_full_upgrade:
        if fa_survey_needed:
            survey_required = "YES"
            survey_type = SURVEY_TYPE_FA_INTRUSION
            upgrade_decision = "FULL CCTV UPGRADE + FA/INTRUSION SURVEY"
            reason = "CCTV full upgrade (coax/siamese cable). FA/Intrusion survey still required."
        else:
            survey_required = "NO"
            survey_type = "NONE"
            upgrade_decision = "FULL CCTV UPGRADE"
            reason = "Coax or Siamese cabling present. Full CCTV upgrade required. No survey needed."
    
    # No full upgrades - check survey triggers
    elif fa_survey_needed and cctv_survey_needed:
        survey_required = "YES"
        survey_type = SURVEY_TYPE_BOTH
        upgrade_decision = "SURVEY REQUIRED"
        reason = "Both FA/Intrusion and CCTV survey triggers present."
    elif fa_survey_needed:
        survey_required = "YES"
        survey_type = SURVEY_TYPE_FA_INTRUSION
        upgrade_decision = "SURVEY REQUIRED"
        reason = "FA/Intrusion survey triggers present."
    elif cctv_survey_needed:
        survey_required = "YES"
        survey_type = SURVEY_TYPE_CCTV
        upgrade_decision = "SURVEY REQUIRED"
        reason = "CCTV survey triggers present."
    else:
        # No survey triggers and no full upgrade triggers - no survey needed
        survey_required = "NO"
        survey_type = "NONE"
        upgrade_decision = "NO TRIGGERS FOUND"
        reason = "Scout data present. No survey or upgrade triggers found."
    
    # === STEP 4: Scheduling Logic ===
    # If survey is already complete, it's not urgent anymore
    if survey_required == "NO":
        schedule_status = "NOT REQUIRED"
    elif survey_complete:
        schedule_status = "COMPLETE"
    elif days is None:
        schedule_status = "REVIEW"
    elif days > CONSTRUCTION_DEADLINE_DAYS:
        schedule_status = "ON TRACK"
    else:
        schedule_status = "URGENT"
    
    # === STEP 5: Ready to Assign ===
    # RO-008 FIX: Scout-completed sites are ready_to_assign=YES
    # The routing decision has been made (we have scout data), so it's ready.
    # Vendor assignment is a separate workflow concern.
    # 
    # Logic:
    # - Scout submitted + survey NOT required = YES (full upgrade path decided)
    # - Scout submitted + survey required = YES (we know what survey type is needed)
    # - Scout NOT submitted = NO (can't make routing decision yet)
    #
    # NOTE: We already returned early if scout is None, so reaching here means scout exists.
    if survey_required in ("YES", "NO"):
        # Scout submitted, routing decision made - ready to assign
        ready_to_assign = "YES"
    else:
        # survey_required is PENDING or REVIEW - needs attention
        ready_to_assign = "NO"
    
    # Handle missing schedule data - affects schedule_status, NOT ready_to_assign
    # Sites without schedule data can still be ready if scout is submitted
    if schedule is None:
        schedule_status = "REVIEW"
    
    # === STEP 6: Vendor Instructions ===
    if survey_type == SURVEY_TYPE_CCTV:
        vendor_instructions = "Complete CCTV mapped survey with GPS. Review supplemental CCTV flags before submission."
    elif survey_type == SURVEY_TYPE_FA_INTRUSION:
        vendor_instructions = "Complete FA/Intrusion mapped survey without GPS. Review supplemental fire alarm flags before submission."
    elif survey_type == SURVEY_TYPE_BOTH:
        vendor_instructions = "Complete CCTV mapped survey with GPS and FA/Intrusion mapped survey without GPS. Review all supplemental flags before submission."
    elif survey_required == "NO":
        vendor_instructions = "No survey required. Site is routed to full upgrade based on scout trigger."
    else:
        vendor_instructions = "Do not assign survey until internal review is completed."
    
    return SurveyRoutingRow(
        site=site,
        vendor=vendor,
        days_to_construction=days_str,
        survey_required=survey_required,
        survey_type=survey_type,
        upgrade_decision=upgrade_decision,
        reason_for_decision=reason,
        schedule_status=schedule_status,
        ready_to_assign=ready_to_assign,
        supplemental_flags="; ".join(supplemental_flags_list),
        vendor_instructions=vendor_instructions,
        survey_complete=survey_complete,
        # Progress tracking from schedule
        assigned=schedule.assigned if schedule else False,
        assigned_vendor=schedule.assigned_vendor if schedule else "",
        survey_returned_qa=schedule.survey_returned_qa if schedule else False,
        passes_at_qa=schedule.passes_at_qa if schedule else False,
        survey_returned_date=schedule.survey_returned_date if schedule else "",
        in_siteowl_design=schedule.in_siteowl_design if schedule else False,
        in_siteowl_installation=schedule.in_siteowl_installation if schedule else False,
        in_siteowl_live=schedule.in_siteowl_live if schedule else False,
        percentage_complete=schedule.percentage_complete if schedule else 0.0,
        on_project_tracking=schedule.on_project_tracking if schedule else False,
        scout_submitted=True,  # Scout was submitted (we have scout data)
    )


def build_survey_routing_data(
    scout_token: str,
    survey_token: str,
    workbook_path: str | Path = DEFAULT_WORKBOOK_PATH,
) -> dict[str, Any]:
    """Build complete survey routing dataset.
    
    Args:
        scout_token: Airtable token for Scout table (appAwgaX89x0JxG3Z)
        survey_token: Airtable token for Survey Submissions (apptK6zNN0Hf3OuoJ)
        workbook_path: Path to Excel workbook with schedule data
    """
    
    # Fetch data from all sources - using appropriate tokens
    scout_records = fetch_scout_data(scout_token)
    schedule_records = load_schedule_data(workbook_path)
    
    # Fetch completed surveys from Airtable Submissions table
    # This catches surveys submitted via Airtable that aren't yet in Excel
    airtable_completed_sites = fetch_survey_submissions(survey_token)
    
    # Index by site
    scout_by_site = {s.site: s for s in scout_records}
    schedule_by_site = {s.site: s for s in schedule_records}
    
    # Get all unique sites
    all_sites = set(scout_by_site.keys()) | set(schedule_by_site.keys())
    
    # Evaluate each site
    rows = []
    for site in sorted(all_sites):
        scout = scout_by_site.get(site)
        schedule = schedule_by_site.get(site)
        row = evaluate_site(scout, schedule)
        row_dict = asdict(row)
        
        # Apply manual scout overrides (sites marked as scout complete with specific survey type)
        if site in MANUAL_SCOUT_COMPLETE:
            row_dict["survey_required"] = "YES"
            row_dict["survey_type"] = MANUAL_SCOUT_COMPLETE[site]
            row_dict["ready_to_assign"] = "YES"
            row_dict["reason_for_decision"] = f"Manual override: Scout complete, {MANUAL_SCOUT_COMPLETE[site]} survey required"
            if row_dict["schedule_status"] == "PENDING":
                row_dict["schedule_status"] = "ON TRACK"
        
        # Override survey_complete if site has PASS in Airtable Submissions
        if site in airtable_completed_sites:
            row_dict["survey_complete"] = True
            if row_dict["schedule_status"] != "NOT REQUIRED":
                row_dict["schedule_status"] = "COMPLETE"
        
        rows.append(row_dict)
    
    # Calculate summary stats
    total = len(rows)
    surveys_required = sum(1 for r in rows if r["survey_required"] == "YES")
    surveys_complete = sum(1 for r in rows if r["survey_complete"])
    full_upgrades = sum(1 for r in rows if "FULL" in r["upgrade_decision"])
    review_required = sum(1 for r in rows if r["survey_required"] == "REVIEW")
    # Urgent = schedule says urgent AND not already complete
    urgent_sites = sum(1 for r in rows if r["schedule_status"] == "URGENT" and not r["survey_complete"])
    ready_to_assign = sum(1 for r in rows if r["ready_to_assign"] == "YES")
    
    # NEW: Pending scout (can't make routing decision)
    pending_scout = sum(1 for r in rows if r["survey_required"] == "PENDING")
    
    # NEW: Sites without vendor assigned
    no_vendor = sum(1 for r in rows if not r["vendor"] and r["survey_required"] == "YES")
    
    # NEW: Sites not on Project Tracking sheet
    not_on_tracking = sum(1 for r in rows if not r["on_project_tracking"])
    
    # NEW: Completed but not on tracking (like site 5027)
    completed_not_listed = sum(1 for r in rows if r["survey_complete"] and not r["on_project_tracking"])
    
    # Survey type breakdown (only for sites that need surveys)
    cctv_surveys = sum(1 for r in rows if r["survey_type"] == SURVEY_TYPE_CCTV)
    fa_surveys = sum(1 for r in rows if r["survey_type"] == SURVEY_TYPE_FA_INTRUSION)
    both_surveys = sum(1 for r in rows if r["survey_type"] == SURVEY_TYPE_BOTH)
    pending_type = sum(1 for r in rows if r["survey_type"] == "PENDING")
    
    # Full upgrade breakdown
    full_cctv = sum(1 for r in rows if "FULL CCTV" in r["upgrade_decision"] or "FULL BOTH" in r["upgrade_decision"])
    full_fa = sum(1 for r in rows if "FULL FA" in r["upgrade_decision"] or "FULL BOTH" in r["upgrade_decision"])
    full_both = sum(1 for r in rows if "FULL BOTH" in r["upgrade_decision"])
    
    # RO-006: Vendor breakdown for auditability
    # Completed sites are excluded from "pending" surveys (they're done!)
    vendor_breakdown = {}
    for r in rows:
        v = r["vendor"] or "unassigned"
        if v not in vendor_breakdown:
            vendor_breakdown[v] = {"total": 0, "survey_required": 0, "pending": 0, "complete": 0}
        vendor_breakdown[v]["total"] += 1
        if r["survey_required"] == "YES":
            vendor_breakdown[v]["survey_required"] += 1
            if r["survey_complete"]:
                vendor_breakdown[v]["complete"] += 1
            else:
                vendor_breakdown[v]["pending"] += 1  # Still needs to be done
    
    return {
        "generated_at": datetime.now().isoformat(),
        "summary": {
            "total_sites": total,
            "surveys_required": surveys_required,
            "surveys_complete": surveys_complete,
            "cctv_surveys": cctv_surveys,
            "fa_surveys": fa_surveys,
            "both_surveys": both_surveys,
            "pending_scout": pending_scout,
            "pending_type": pending_type,
            "full_upgrades": full_upgrades,
            "full_cctv": full_cctv,
            "full_fa": full_fa,
            "full_both": full_both,
            "review_required": review_required,
            "urgent_sites": urgent_sites,
            "ready_to_assign": ready_to_assign,
            "no_vendor": no_vendor,
            "not_on_tracking": not_on_tracking,
            "completed_not_listed": completed_not_listed,
            "vendor_breakdown": vendor_breakdown,
        },
        "rows": rows,
    }


def refresh_survey_routing(
    scout_token: str,
    survey_token: str,
    output_dir: Path | str,
    workbook_path: str | Path = DEFAULT_WORKBOOK_PATH,
    sync_to_airtable: bool = True
) -> None:
    """Refresh survey routing data and write to JSON.
    
    Args:
        scout_token: Airtable token for Scout table
        survey_token: Airtable token for Survey Submissions
        output_dir: Directory to write JSON output
        workbook_path: Path to Excel workbook with MAP DATA
        sync_to_airtable: If True, also sync updates to Airtable Survey Routing table
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    data = build_survey_routing_data(scout_token, survey_token, workbook_path)
    
    output_file = output_dir / "survey_routing_data.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    
    log.info(
        f"Survey routing data updated: {data['summary']['total_sites']} sites, "
        f"{data['summary']['surveys_required']} surveys required, "
        f"{data['summary']['full_upgrades']} full upgrades"
    )
    
    # Sync to Airtable Survey Routing table (uses scout token - same base)
    if sync_to_airtable:
        try:
            updated, errors = sync_routing_to_airtable(scout_token, data["rows"])
            log.info(f"Airtable Survey Routing sync: {updated} updated, {errors} errors")
        except Exception as e:
            log.error(f"Failed to sync to Airtable Survey Routing: {e}")


def _derive_status(row: dict) -> str:
    """Derive display status from routing row data."""
    if row.get('survey_complete') is True:
        return 'Completed'
    
    schedule = row.get('schedule_status', '')
    ready = row.get('ready_to_assign', '')
    survey_req = row.get('survey_required', '')
    reason = row.get('reason_for_decision', '')
    
    if survey_req == 'NO':
        return 'No Survey Needed'
    if 'No scout submission' in str(reason):
        return 'Awaiting Scout'
    if schedule == 'REVIEW':
        return 'Needs Review'
    elif schedule == 'ON TRACK':
        if ready == 'YES':
            return 'Ready to Assign'
        else:
            return 'In Progress'
    elif schedule == 'NOT REQUIRED':
        return 'No Survey Needed'
    
    return 'Pending'


def sync_routing_to_airtable(token: str, rows: list[dict]) -> tuple[int, int]:
    """Sync routing data to Airtable Survey Routing table.
    
    Args:
        token: Airtable API token
        rows: List of routing row dicts from build_survey_routing_data
        
    Returns:
        Tuple of (updated_count, error_count)
    """
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }
    
    # Build lookup by site
    site_data = {str(r.get('site', '')).strip(): r for r in rows}
    
    # Fetch all current Airtable records
    all_records = []
    offset = None
    
    while True:
        url = f'https://api.airtable.com/v0/{SCOUT_BASE_ID}/{ROUTING_TABLE_ID}'
        params = {'pageSize': 100}
        if offset:
            params['offset'] = offset
        
        try:
            resp = requests.get(url, headers=headers, params=params, timeout=30)
            resp.raise_for_status()
        except Exception as e:
            log.error(f"Failed to fetch Survey Routing records: {e}")
            return (0, 0)
        
        result = resp.json()
        all_records.extend(result.get('records', []))
        offset = result.get('offset')
        if not offset:
            break
    
    log.info(f"Fetched {len(all_records)} records from Survey Routing table")
    
    # Build list of updates needed
    updates = []
    for rec in all_records:
        site = str(rec['fields'].get('Site', '')).strip()
        row = site_data.get(site)
        
        if not row:
            continue
        
        new_status = _derive_status(row)
        current_status = rec['fields'].get('Status', '')
        
        # Only update if status changed or survey type changed
        new_survey_type = row.get('survey_type', '')
        current_survey_type = rec['fields'].get('Survey Type', '')
        
        if new_status != current_status or new_survey_type != current_survey_type:
            updates.append({
                'id': rec['id'],
                'fields': {
                    'Status': new_status,
                    'Survey Type': new_survey_type,
                    'Notes': row.get('notes', '') or row.get('vendor_instructions', ''),
                }
            })
    
    if not updates:
        log.info("No Survey Routing updates needed")
        return (0, 0)
    
    log.info(f"Updating {len(updates)} Survey Routing records...")
    
    # Update in batches of 10 (Airtable limit)
    import time
    BATCH_SIZE = 10
    updated = 0
    errors = 0
    
    for i in range(0, len(updates), BATCH_SIZE):
        batch = updates[i:i+BATCH_SIZE]
        
        try:
            resp = requests.patch(
                f'https://api.airtable.com/v0/{SCOUT_BASE_ID}/{ROUTING_TABLE_ID}',
                headers=headers,
                json={'records': batch},
                timeout=30
            )
            if resp.ok:
                updated += len(batch)
            else:
                errors += len(batch)
                log.error(f"Batch update failed: {resp.text[:200]}")
        except Exception as e:
            errors += len(batch)
            log.error(f"Batch update exception: {e}")
        
        # Rate limiting: max 5 requests per second
        time.sleep(0.25)
    
    return (updated, errors)
