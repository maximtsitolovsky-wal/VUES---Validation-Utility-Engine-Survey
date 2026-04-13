import sys
import openpyxl

xl_path = r'C:\Users\vn59j7j\OneDrive - Walmart Inc\7162 Test.xlsx'
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(xl_path, data_only=True)
ws = wb.active
headers = [cell.value for cell in ws[1]]

grading_cols = {'Name', 'Abbreviated Name', 'Part Number', 'Manufacturer',
                'IP Address', 'MAC Address', 'IP / Analog', 'Description',
                'Project ID', 'Plan ID'}

print(f"Total rows: {ws.max_row}, Total cols: {ws.max_column}")
print()

# Scan ALL rows, flag any blank grading columns
blank_issues = []
for row_num in range(2, ws.max_row + 1):
    row = list(ws[row_num])
    for i, cell in enumerate(row):
        col_name = headers[i]
        if col_name in grading_cols and col_name != 'Project ID' and col_name != 'Plan ID':
            if cell.value is None or str(cell.value).strip() == '':
                blank_issues.append((row_num, col_name))

if blank_issues:
    print(f"BLANK GRADING COLUMNS ({len(blank_issues)} found):")
    for row_num, col in blank_issues[:30]:
        print(f"  Row {row_num}: [{col}] is EMPTY")
    if len(blank_issues) > 30:
        print(f"  ... and {len(blank_issues) - 30} more")
else:
    print("No blank grading columns found.")

print()
print("=== ROWS 44-48 DETAIL ===")
for row_num in range(44, 49):
    if row_num > ws.max_row:
        break
    row = list(ws[row_num])
    print(f"\nROW {row_num}:")
    for i, cell in enumerate(row):
        if headers[i] in grading_cols:
            val = repr(cell.value)
            print(f"  [{headers[i]:25}]: {val}")
