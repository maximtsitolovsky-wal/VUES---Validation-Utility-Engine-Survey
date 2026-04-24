import openpyxl
import json

# Load Everon assignments
wb = openpyxl.load_workbook(r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\Excel\Vendor ASSIGN. 4.2.26.xlsx', data_only=True)
ws = wb['Everon']

everon_sites = set()
for row in range(1, ws.max_row + 1):
    for col in range(1, 10):
        cell = ws.cell(row, col).value
        if cell:
            try:
                everon_sites.add(str(int(cell)))
            except:
                pass

print(f"Everon assigned sites: {len(everon_sites)}")
print(f"Sample: {list(everon_sites)[:10]}")

# Load Scout records
with open('output/team_dashboard_data.json') as f:
    data = json.load(f)

# Find overlaps
matches = []
for rec in data['scout']['records']:
    rf = rec.get('raw_fields', {})
    site = str(rf.get('Site Number', '')).strip()
    try:
        site = str(int(float(site)))
    except:
        pass
    
    if site in everon_sites:
        vendor = rf.get('Surveyor Parent Company', 'Unknown')
        complete = rf.get('Complete?')
        matches.append((site, vendor, complete))

print(f"\nEveron-assigned sites found in Scout: {len(matches)}")
for m in matches:
    print(f"  Site {m[0]}: vendor={m[1]}, complete={m[2]}")
