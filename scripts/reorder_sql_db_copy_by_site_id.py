from __future__ import annotations

import json
import sqlite3
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet._read_only import ReadOnlyCell

SOURCE_PATH = Path(r"C:\Users\vn59j7j\OneDrive - Walmart Inc\SQL DB COPY.xlsx")
OUTPUT_SUFFIX = "__siteid_grouped_stable.xlsx"
SITE_ID_HEADER = "Site ID"
VALIDATION_SUFFIX = "__siteid_grouped_stable.validation.json"


CellLike = Cell | ReadOnlyCell | None


def safe_header(value: object) -> str:
    return "" if value is None else str(value)



def should_drop_header(header: str) -> bool:
    return header.startswith("Column")



def cell_value(cell: CellLike):
    return None if cell is None else cell.value



def build_output_path(source_path: Path) -> Path:
    return source_path.with_name(f"{source_path.stem}{OUTPUT_SUFFIX}")



def build_validation_path(output_path: Path) -> Path:
    return output_path.with_name(f"{output_path.stem}.validation.json")



def main() -> None:
    if not SOURCE_PATH.exists():
        raise FileNotFoundError(f"Source workbook not found: {SOURCE_PATH}")

    output_path = build_output_path(SOURCE_PATH)
    validation_path = build_validation_path(output_path)

    with tempfile.TemporaryDirectory(prefix="siteid_reorder_") as temp_dir:
        db_path = Path(temp_dir) / "rows.sqlite"
        stats = reorder_workbook(SOURCE_PATH, output_path, db_path)
        validation_path.write_text(json.dumps(stats, indent=2), encoding="utf-8")

    print(f"OUTPUT_FILE={output_path}")
    print(f"VALIDATION_FILE={validation_path}")
    print(f"SHEET_NAME={stats['sheet_name']}")
    print(f"SOURCE_ROW_COUNT={stats['source_row_count']}")
    print(f"OUTPUT_ROW_COUNT={stats['output_row_count']}")
    print(f"REMOVED_COLUMN_HEADERS={stats['removed_column_header_count']}")
    print(f"KEPT_COLUMN_HEADERS={stats['kept_column_count']}")
    print(f"SITE_ID_GROUP_COUNT={stats['site_id_group_count']}")



def reorder_workbook(source_path: Path, output_path: Path, db_path: Path) -> dict[str, object]:
    wb = load_workbook(source_path, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    sheet_name = ws.title

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    keep_indexes: list[int] = []
    kept_headers: list[str] = []
    removed_headers: list[str] = []
    site_id_index: int | None = None

    for idx, cell in enumerate(header_cells):
        header = safe_header(cell_value(cell))
        if header == "":
            continue
        if should_drop_header(header):
            removed_headers.append(header)
            continue
        keep_indexes.append(idx)
        kept_headers.append(header)
        if header == SITE_ID_HEADER:
            site_id_index = idx

    if site_id_index is None:
        raise ValueError(f"Required header not found: {SITE_ID_HEADER}")

    conn = sqlite3.connect(db_path)
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=OFF")
        conn.execute("PRAGMA temp_store=MEMORY")
        conn.execute(
            "CREATE TABLE rows (orig_idx INTEGER PRIMARY KEY, site_id TEXT NOT NULL, payload TEXT NOT NULL)"
        )

        source_row_count = 0
        insert_batch: list[tuple[int, str, str]] = []
        batch_size = 5000

        for orig_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
            source_row_count += 1
            site_id = safe_header(cell_value(row[site_id_index]))
            kept_values = [cell_value(row[i]) for i in keep_indexes]
            insert_batch.append((orig_idx, site_id, json.dumps(kept_values, ensure_ascii=False)))
            if len(insert_batch) >= batch_size:
                conn.executemany(
                    "INSERT INTO rows (orig_idx, site_id, payload) VALUES (?, ?, ?)",
                    insert_batch,
                )
                conn.commit()
                insert_batch.clear()

        if insert_batch:
            conn.executemany(
                "INSERT INTO rows (orig_idx, site_id, payload) VALUES (?, ?, ?)",
                insert_batch,
            )
            conn.commit()

        conn.execute("CREATE INDEX idx_rows_site_orig ON rows (site_id, orig_idx)")
        conn.commit()

        out_wb = Workbook(write_only=True)
        out_ws = out_wb.create_sheet(title=sheet_name)
        out_ws.append(kept_headers)

        output_row_count = 0
        cursor = conn.execute(
            "SELECT payload FROM rows ORDER BY site_id, orig_idx"
        )
        for (payload,) in cursor:
            out_ws.append(json.loads(payload))
            output_row_count += 1

        if "Sheet" in out_wb.sheetnames:
            std = out_wb["Sheet"]
            out_wb.remove(std)

        out_wb.save(output_path)

        site_id_group_count = conn.execute(
            "SELECT COUNT(DISTINCT site_id) FROM rows"
        ).fetchone()[0]
    finally:
        conn.close()
        wb.close()

    return {
        "source_file": str(source_path),
        "output_file": str(output_path),
        "sheet_name": sheet_name,
        "source_row_count": source_row_count,
        "output_row_count": output_row_count,
        "removed_column_header_count": len(removed_headers),
        "removed_column_headers_sample": removed_headers[:25],
        "kept_column_count": len(kept_headers),
        "kept_headers": kept_headers,
        "site_id_group_count": site_id_group_count,
        "ordering_rule": "ORDER BY Site ID, original row index",
    }


if __name__ == "__main__":
    main()
