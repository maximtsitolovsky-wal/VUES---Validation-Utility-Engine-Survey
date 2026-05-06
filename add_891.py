#!/usr/bin/env python3
"""Add site 891 to Scout Map Data."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from pathlib import Path

scout_file = Path(r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\ScoutSurveyLab.xlsm')

# First check what columns exist
print("Loading ScoutSurveyLab.xlsm...")
wb = openpyxl.load_workbook(scout_file, keep_vba=True)  # keep_vba for .xlsm
ws = wb['Scout Map Data']

# Get headers
headers = [cell.value for cell in ws[1]]
print(f"Headers: {headers[:10]}...")

# Find last row
last_row = ws.max_row
print(f"Current rows: {last_row}")

# Check if 891 already exists (double-check)
for row in ws.iter_rows(min_row=2, max_col=1):
    if row[0].value and str(row[0].value).strip().lstrip('0') == '891':
        print("Site 891 already exists! No action needed.")
        wb.close()
        sys.exit(0)

# Add 891 to the next row
new_row = last_row + 1
ws.cell(row=new_row, column=1, value=891)  # Site number in column A
print(f"Added site 891 at row {new_row}")

# Save
wb.save(scout_file)
wb.close()
print(f"✅ Saved {scout_file.name}")

# Verify
wb2 = openpyxl.load_workbook(scout_file, read_only=True, data_only=True)
ws2 = wb2['Scout Map Data']
count = sum(1 for row in ws2.iter_rows(min_row=2, max_col=1) if row[0].value)
wb2.close()
print(f"Scout Map Data now has {count} sites")
