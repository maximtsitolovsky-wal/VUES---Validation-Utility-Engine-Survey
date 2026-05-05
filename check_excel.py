import openpyxl
from pathlib import Path

excel_path = Path(r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\Excel\Vendor ASSIGN. 4.2.26.xlsx')
print(f"Looking for: {excel_path}")
print(f"Exists: {excel_path.exists()}")

if excel_path.exists():
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    print(f"Sheets: {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    print(f"Columns: {headers}")
    
    # Get first 2 rows of data
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4)):
        print(f"Row {i+2}: {[cell.value for cell in row[:12]]}")
    wb.close()
else:
    # Try to find any Excel files
    print("\nSearching for Excel files...")
    for p in Path(r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents').rglob('*.xlsx'):
        print(f"  Found: {p}")
