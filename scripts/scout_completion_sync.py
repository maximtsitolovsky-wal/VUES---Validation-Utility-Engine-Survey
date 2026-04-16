"""
Scout Completion Sync — Airtable → Excel Status Tracker
========================================================
Syncs completion status from Scout Airtable into ScoutSurveyLab.xlsm

Mapping:
  Airtable "Site Number"    → Excel "Store Number"
  Airtable "Complete?"      → Excel "Completed Scout" (True/False)
  When marking complete     → Excel "Confirmed by" = "code puppy"
                            → Excel date/time stamp

Schedule: Mon-Fri 10:00 AM and 3:00 PM via Windows Task Scheduler.
Register:  ops/windows/register_scout_completion_sync_task.ps1
"""
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────
API_KEY   = "patPR0WWxXCE0loRO.d18126548ad25b8aaf9fd43e2ac69479b1378e46d7f8c6efbdd88f7197a4d495"
BASE_ID   = "appAwgaX89x0JxG3Z"
TABLE_ID  = "tblC4o9AvVulyxFMk"

AIRTABLE_SITE_COL = "Site Number"
AIRTABLE_COMPLETE_COL = "Complete?"
AIRTABLE_VENDOR_COL = "Vendor Parent Company"
AIRTABLE_SURVEYOR_COL = "Surveyor Name"

EXCEL_PATH = Path(r"C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\ScoutSurveyLab.xlsm")
EXCEL_SHEET_NAME = "Scout Map Data"
EXCEL_OUTLIER_SHEET_NAME = "Outlier Scout"
EXCEL_STORE_COL = "Store Number"
EXCEL_COMPLETED_COL = "Completed Scout"
EXCEL_CONFIRMED_BY_COL = "Confirmed by"

LOG_FILE = Path(__file__).resolve().parents[1] / "logs" / "scout_completion_sync.log"

AIRTABLE_URL = f"https://api.airtable.com/v0/{BASE_ID}/{TABLE_ID}"
AIRTABLE_HDR = {"Authorization": f"Bearer {API_KEY}"}


# ── Helpers ───────────────────────────────────────────────────────────────────
def log(msg: str) -> None:
    """Log to console and file."""
    print(msg, flush=True)
    try:
        LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
        with LOG_FILE.open("a", encoding="utf-8") as f:
            f.write(f"{datetime.now().isoformat()} {msg}\n")
    except Exception:
        pass


def fetch_airtable_records() -> list[dict]:
    """Fetch all records from Scout Airtable with pagination."""
    records, params = [], {}
    while True:
        resp = requests.get(AIRTABLE_URL, headers=AIRTABLE_HDR, params=params, timeout=30)
        resp.raise_for_status()
        body = resp.json()
        records.extend(body.get("records", []))
        offset = body.get("offset")
        if not offset:
            break
        params = {"offset": offset}
        time.sleep(0.2)
    return records


def parse_site_number(value) -> str:
    """Normalize site/store number to string for matching."""
    if value is None:
        return ""
    return str(value).strip()


def is_complete(complete_value) -> bool:
    """Check if Airtable 'Complete?' field is truthy."""
    if complete_value is None:
        return False
    if isinstance(complete_value, bool):
        return complete_value
    if isinstance(complete_value, str):
        return complete_value.lower() in ("true", "yes", "1", "✓", "x")
    return bool(complete_value)


def sync_completion_status() -> tuple[int, int, int]:
    """
    Sync completion status from Airtable to Excel.
    
    Returns:
        (updated, skipped, errors) counts
    """
    log("[*] Fetching records from Airtable...")
    airtable_records = fetch_airtable_records()
    log(f"[OK] Fetched {len(airtable_records)} records from Airtable")

    # Build completion map: site_number -> {is_complete, vendor, surveyor}
    completion_map = {}
    for rec in airtable_records:
        fields = rec.get("fields", {})
        site_num = parse_site_number(fields.get(AIRTABLE_SITE_COL))
        complete = is_complete(fields.get(AIRTABLE_COMPLETE_COL))
        vendor = fields.get(AIRTABLE_VENDOR_COL, "")
        surveyor = fields.get(AIRTABLE_SURVEYOR_COL, "")
        if site_num:
            completion_map[site_num] = {
                "complete": complete,
                "vendor": vendor,
                "surveyor": surveyor
            }

    log(f"[*] Built completion map with {len(completion_map)} sites")
    log(f"[*] Sites marked complete: {sum(1 for v in completion_map.values() if v['complete'])}")

    # Open Excel workbook
    if not EXCEL_PATH.exists():
        log(f"[ERROR] Excel file not found: {EXCEL_PATH}")
        return 0, 0, 1

    log(f"[*] Opening Excel: {EXCEL_PATH.name}...")
    
    # Try to open with retry logic (file might be locked by Excel/OneDrive)
    wb = None
    max_retries = 3
    for attempt in range(1, max_retries + 1):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
            break
        except PermissionError as e:
            if attempt < max_retries:
                log(f"[WARN] File locked (attempt {attempt}/{max_retries}). Retrying in 2s...")
                log(f"       Please close Excel if {EXCEL_PATH.name} is open.")
                time.sleep(2)
            else:
                log(f"[ERROR] File locked after {max_retries} attempts.")
                log(f"        Please close {EXCEL_PATH.name} in Excel and try again.")
                raise e
    
    if wb is None:
        log("[ERROR] Failed to open Excel workbook")
        return 0, 0, 1
    
    # Get the Scout Map Data sheet
    if EXCEL_SHEET_NAME not in wb.sheetnames:
        log(f"[ERROR] Sheet '{EXCEL_SHEET_NAME}' not found in workbook")
        log(f"        Available sheets: {', '.join(wb.sheetnames)}")
        wb.close()
        return 0, 0, 1
    
    ws = wb[EXCEL_SHEET_NAME]
    log(f"[*] Using sheet: {ws.title}")
    
    # Get or create Outlier Scout sheet
    if EXCEL_OUTLIER_SHEET_NAME not in wb.sheetnames:
        log(f"[*] Creating new sheet: {EXCEL_OUTLIER_SHEET_NAME}")
        outlier_ws = wb.create_sheet(EXCEL_OUTLIER_SHEET_NAME)
        # Add headers
        outlier_ws.cell(1, 1, "Site Number")
        outlier_ws.cell(1, 2, "Vendor Parent Company")
        outlier_ws.cell(1, 3, "Surveyor Name")
    else:
        outlier_ws = wb[EXCEL_OUTLIER_SHEET_NAME]
        log(f"[*] Found existing sheet: {EXCEL_OUTLIER_SHEET_NAME}")

    # Find header row and column indices
    header_row = None
    store_col_idx = None
    completed_col_idx = None
    confirmed_by_col_idx = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if cell.value:
                cell_val = str(cell.value).strip()
                if cell_val == EXCEL_STORE_COL:
                    header_row = row_idx
                    store_col_idx = col_idx
                elif cell_val == EXCEL_COMPLETED_COL:
                    completed_col_idx = col_idx
                elif cell_val == EXCEL_CONFIRMED_BY_COL:
                    confirmed_by_col_idx = col_idx
        
        if header_row and store_col_idx and completed_col_idx:
            break

    if not all([header_row, store_col_idx, completed_col_idx]):
        log(f"[ERROR] Could not find required columns. Found: Store={store_col_idx}, Completed={completed_col_idx}, Header Row={header_row}")
        wb.close()
        return 0, 0, 1

    log(f"[OK] Found headers at row {header_row}: Store col={store_col_idx}, Completed col={completed_col_idx}, Confirmed by col={confirmed_by_col_idx}")

    # Build set of existing store numbers in Scout Map Data
    existing_stores = set()
    for row_idx in range(header_row + 1, ws.max_row + 1):
        store_num_cell = ws.cell(row_idx, store_col_idx)
        store_num = parse_site_number(store_num_cell.value)
        if store_num:
            existing_stores.add(store_num)
    
    log(f"[*] Found {len(existing_stores)} existing stores in '{EXCEL_SHEET_NAME}'")

    # Track outliers (sites in Airtable but not in Scout Map Data)
    outliers = []
    for site_num, data in completion_map.items():
        if site_num not in existing_stores:
            outliers.append({
                "site_number": site_num,
                "vendor": data["vendor"],
                "surveyor": data["surveyor"]
            })
    
    if outliers:
        log(f"[*] Found {len(outliers)} outlier site(s) not in Scout Map Data")
    
    # Process data rows
    updated = 0
    skipped = 0
    errors = 0
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row_idx in range(header_row + 1, ws.max_row + 1):
        store_num_cell = ws.cell(row_idx, store_col_idx)
        store_num = parse_site_number(store_num_cell.value)
        
        if not store_num:
            continue  # Skip empty rows

        # Check if this site is in our completion map
        if store_num not in completion_map:
            skipped += 1
            continue

        site_data = completion_map[store_num]
        airtable_complete = site_data["complete"]
        completed_cell = ws.cell(row_idx, completed_col_idx)
        current_value = completed_cell.value

        # Check if current value is already True/complete
        current_complete = False
        if isinstance(current_value, bool):
            current_complete = current_value
        elif isinstance(current_value, str):
            current_complete = current_value.lower() in ("true", "yes", "1", "✓", "x")

        # Update if Airtable says complete but Excel doesn't
        if airtable_complete and not current_complete:
            completed_cell.value = True
            
            # Update "Confirmed by" column if it exists
            if confirmed_by_col_idx:
                confirmed_by_cell = ws.cell(row_idx, confirmed_by_col_idx)
                confirmed_by_cell.value = f"code puppy ({now_str})"
            
            log(f"   [UPDATE] Store {store_num}: Completed Scout -> True")
            updated += 1
        else:
            skipped += 1
    
    # Write outliers to Outlier Scout sheet
    outliers_added = 0
    if outliers:
        # Get existing outliers to avoid duplicates
        existing_outliers = set()
        for row_idx in range(2, outlier_ws.max_row + 1):
            site = parse_site_number(outlier_ws.cell(row_idx, 1).value)
            if site:
                existing_outliers.add(site)
        
        # Add new outliers
        next_row = outlier_ws.max_row + 1
        for outlier in outliers:
            if outlier["site_number"] not in existing_outliers:
                outlier_ws.cell(next_row, 1, outlier["site_number"])
                outlier_ws.cell(next_row, 2, outlier["vendor"])
                outlier_ws.cell(next_row, 3, outlier["surveyor"])
                log(f"   [OUTLIER] Added site {outlier['site_number']} to Outlier Scout")
                next_row += 1
                outliers_added += 1

    # Save workbook if changes were made
    if updated > 0 or outliers_added > 0:
        log(f"[*] Saving changes to Excel...")
        
        # Try to save with retry logic (OneDrive might be syncing)
        save_retries = 3
        saved = False
        for attempt in range(1, save_retries + 1):
            try:
                wb.save(EXCEL_PATH)
                saved = True
                log(f"[OK] Saved {updated} updates + {outliers_added} outliers to {EXCEL_PATH.name}")
                break
            except PermissionError:
                if attempt < save_retries:
                    log(f"[WARN] Save blocked (attempt {attempt}/{save_retries}). Retrying in 2s...")
                    time.sleep(2)
                else:
                    log(f"[ERROR] Failed to save after {save_retries} attempts.")
                    log(f"        OneDrive or Excel may have locked the file.")
                    raise
        
        if not saved:
            wb.close()
            return 0, skipped, 1
    else:
        log("[*] No updates needed - Excel already in sync")

    wb.close()
    return updated, skipped, errors


# ── Main ──────────────────────────────────────────────────────────────────────
def main() -> None:
    log("=" * 80)
    log(f"[START] Scout Completion Sync — {datetime.now().isoformat()}")
    log("=" * 80)

    try:
        updated, skipped, errors = sync_completion_status()
        log("\n" + "=" * 80)
        log(f"[DONE] Updated={updated}  Skipped={skipped}  Errors={errors}")
        log("=" * 80)
        
        sys.exit(1 if errors > 0 else 0)
    
    except Exception as exc:
        log(f"[FATAL] {exc}")
        import traceback
        log(traceback.format_exc())
        sys.exit(1)


if __name__ == "__main__":
    main()
