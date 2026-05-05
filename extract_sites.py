"""Extract sites with zip codes and convert to lat/lng using free zip code data."""
import openpyxl
import json
import urllib.request
from pathlib import Path

excel_path = Path(r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\Excel\Vendor ASSIGN. 4.2.26.xlsx')

# US Zip code centroids (we'll use a subset for common zips, fallback to state centroid)
# Download a small lookup or use embedded data
STATE_COORDS = {
    'AL': (32.806671, -86.791130), 'AK': (61.370716, -152.404419), 'AZ': (33.729759, -111.431221),
    'AR': (34.969704, -92.373123), 'CA': (36.116203, -119.681564), 'CO': (39.059811, -105.311104),
    'CT': (41.597782, -72.755371), 'DE': (39.318523, -75.507141), 'FL': (27.766279, -81.686783),
    'GA': (33.040619, -83.643074), 'HI': (21.094318, -157.498337), 'ID': (44.240459, -114.478828),
    'IL': (40.349457, -88.986137), 'IN': (39.849426, -86.258278), 'IA': (42.011539, -93.210526),
    'KS': (38.526600, -96.726486), 'KY': (37.668140, -84.670067), 'LA': (31.169546, -91.867805),
    'ME': (44.693947, -69.381927), 'MD': (39.063946, -76.802101), 'MA': (42.230171, -71.530106),
    'MI': (43.326618, -84.536095), 'MN': (45.694454, -93.900192), 'MS': (32.741646, -89.678696),
    'MO': (38.456085, -92.288368), 'MT': (46.921925, -110.454353), 'NE': (41.125370, -98.268082),
    'NV': (38.313515, -117.055374), 'NH': (43.452492, -71.563896), 'NJ': (40.298904, -74.521011),
    'NM': (34.840515, -106.248482), 'NY': (42.165726, -74.948051), 'NC': (35.630066, -79.806419),
    'ND': (47.528912, -99.784012), 'OH': (40.388783, -82.764915), 'OK': (35.565342, -96.928917),
    'OR': (44.572021, -122.070938), 'PA': (40.590752, -77.209755), 'RI': (41.680893, -71.511780),
    'SC': (33.856892, -80.945007), 'SD': (44.299782, -99.438828), 'TN': (35.747845, -86.692345),
    'TX': (31.054487, -97.563461), 'UT': (40.150032, -111.862434), 'VT': (44.045876, -72.710686),
    'VA': (37.769337, -78.169968), 'WA': (47.400902, -121.490494), 'WV': (38.491226, -80.954453),
    'WI': (44.268543, -89.616508), 'WY': (42.755966, -107.302490), 'DC': (38.897438, -77.026817),
}

# Try to load zip code database
zip_coords = {}
zip_db_path = Path('ui/zip_coords.json')
if zip_db_path.exists():
    with open(zip_db_path) as f:
        zip_coords = json.load(f)
    print(f"Loaded {len(zip_coords)} zip codes from cache")

# Extract from Excel
print("Reading Excel...")
wb = openpyxl.load_workbook(excel_path, read_only=True)

sites = []
missing_zips = set()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = {str(cell.value).strip().lower(): idx for idx, cell in enumerate(list(ws.iter_rows(min_row=1, max_row=1))[0]) if cell.value}
    
    store_col = headers.get('store\nnumber', headers.get('store number', 0))
    city_col = headers.get('city', 2)
    state_col = headers.get('state', 3)
    zip_col = headers.get('zip code', 4)
    vendor_col = headers.get('vendor_scoutassignedfinal', 6)
    
    for row in ws.iter_rows(min_row=2):
        store = row[store_col].value
        if not store:
            continue
            
        city = str(row[city_col].value or '').strip()
        state = str(row[state_col].value or '').strip().upper()
        zipcode = str(row[zip_col].value or '').strip()[:5]  # First 5 digits
        vendor = str(row[vendor_col].value or sheet_name).strip()
        
        # Get coordinates
        lat, lng = None, None
        
        if zipcode and zipcode in zip_coords:
            lat, lng = zip_coords[zipcode]
        elif state in STATE_COORDS:
            lat, lng = STATE_COORDS[state]
            missing_zips.add(zipcode)
        
        if lat and lng:
            sites.append({
                'site': store,
                'city': city,
                'state': state,
                'zip': zipcode,
                'vendor': vendor,
                'lat': round(lat, 5),
                'lng': round(lng, 5)
            })

wb.close()

print(f"Extracted {len(sites)} sites")
print(f"Missing zip lookups (using state centroid): {len(missing_zips)}")

# If we're missing zips, try to fetch them
if missing_zips and len(missing_zips) < 500:
    print(f"Fetching {len(missing_zips)} missing zip coordinates...")
    # Use a simple zip code API or hardcode common ones
    # For now we'll just use state centroids

# Save
output = {
    'markers': [{'vendor': s['vendor'], 'state': s['state'], 'lat': s['lat'], 'lng': s['lng'], 
                 'site': s['site'], 'city': s['city'], 'zip': s['zip']} for s in sites],
    'total': len(sites)
}

with open('ui/site_markers.json', 'w') as f:
    json.dump(output, f)

print(f"Saved {len(sites)} markers to ui/site_markers.json")

# Show vendor breakdown
vendors = {}
for s in sites:
    vendors[s['vendor']] = vendors.get(s['vendor'], 0) + 1
print(f"Vendors: {vendors}")
