"""Extract vendor assignment locations for globe visualization."""
import json
import openpyxl
from pathlib import Path
from collections import defaultdict

# US State centroids (approximate lat/lng)
STATE_COORDS = {
    'AL': [32.806671, -86.791130], 'AK': [61.370716, -152.404419], 'AZ': [33.729759, -111.431221],
    'AR': [34.969704, -92.373123], 'CA': [36.116203, -119.681564], 'CO': [39.059811, -105.311104],
    'CT': [41.597782, -72.755371], 'DE': [39.318523, -75.507141], 'FL': [27.766279, -81.686783],
    'GA': [33.040619, -83.643074], 'HI': [21.094318, -157.498337], 'ID': [44.240459, -114.478828],
    'IL': [40.349457, -88.986137], 'IN': [39.849426, -86.258278], 'IA': [42.011539, -93.210526],
    'KS': [38.526600, -96.726486], 'KY': [37.668140, -84.670067], 'LA': [31.169546, -91.867805],
    'ME': [44.693947, -69.381927], 'MD': [39.063946, -76.802101], 'MA': [42.230171, -71.530106],
    'MI': [43.326618, -84.536095], 'MN': [45.694454, -93.900192], 'MS': [32.741646, -89.678696],
    'MO': [38.456085, -92.288368], 'MT': [46.921925, -110.454353], 'NE': [41.125370, -98.268082],
    'NV': [38.313515, -117.055374], 'NH': [43.452492, -71.563896], 'NJ': [40.298904, -74.521011],
    'NM': [34.840515, -106.248482], 'NY': [42.165726, -74.948051], 'NC': [35.630066, -79.806419],
    'ND': [47.528912, -99.784012], 'OH': [40.388783, -82.764915], 'OK': [35.565342, -96.928917],
    'OR': [44.572021, -122.070938], 'PA': [40.590752, -77.209755], 'RI': [41.680893, -71.511780],
    'SC': [33.856892, -80.945007], 'SD': [44.299782, -99.438828], 'TN': [35.747845, -86.692345],
    'TX': [31.054487, -97.563461], 'UT': [40.150032, -111.862434], 'VT': [44.045876, -72.710686],
    'VA': [37.769337, -78.169968], 'WA': [47.400902, -121.490494], 'WV': [38.491226, -80.954453],
    'WI': [44.268543, -89.616508], 'WY': [42.755966, -107.302490], 'DC': [38.897438, -77.026817],
    'PR': [18.220833, -66.590149]
}

# Vendor colors (matching UI)
VENDOR_COLORS = {
    'CEI': '#2ed7ff',       # cyan
    'Wachter': '#a778ff',   # purple
    'Techwise': '#ff5fd2',  # pink
    'Everon': '#24e6a0',    # mint
    'SAS': '#ffc857',       # yellow
}

excel_path = Path(r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\Excel\Vendor ASSIGN. 4.2.26.xlsx')

def extract_locations():
    """Extract vendor assignment locations from Excel."""
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    
    locations = []
    state_vendor_counts = defaultdict(lambda: defaultdict(int))
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        vendor_name = sheet_name
        
        # Find State column
        headers = {cell.value: idx for idx, cell in enumerate(list(ws.iter_rows(min_row=1, max_row=1))[0])}
        state_col = headers.get('State', 3)  # Default to column 4 (0-indexed 3)
        
        for row in ws.iter_rows(min_row=2):
            state = str(row[state_col].value or '').strip().upper()
            if state and state in STATE_COORDS:
                state_vendor_counts[state][vendor_name] += 1
    
    wb.close()
    
    # Create markers with jitter for overlapping states
    markers = []
    for state, vendors in state_vendor_counts.items():
        base_lat, base_lng = STATE_COORDS[state]
        
        # Create marker for each vendor in this state
        offset = 0
        for vendor, count in vendors.items():
            # Add small jitter to separate overlapping markers
            lat = base_lat + (offset * 0.5)
            lng = base_lng + (offset * 0.3)
            offset += 1
            
            markers.append({
                'id': f'{state}-{vendor}',
                'vendor': vendor,
                'state': state,
                'location': [lat, lng],
                'count': count,
                'color': VENDOR_COLORS.get(vendor, '#ffffff')
            })
    
    return {
        'markers': markers,
        'state_totals': {state: dict(vendors) for state, vendors in state_vendor_counts.items()},
        'vendor_totals': {v: sum(state_vendor_counts[s].get(v, 0) for s in state_vendor_counts) for v in VENDOR_COLORS}
    }

if __name__ == '__main__':
    data = extract_locations()
    
    # Save to JSON for the globe visualization
    output_path = Path('output/vendor_locations.json')
    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)
    
    print(f'Extracted {len(data["markers"])} markers across {len(data["state_totals"])} states')
    print(f'Vendor totals: {data["vendor_totals"]}')
    
    # Copy to UI
    import shutil
    shutil.copy(output_path, 'ui/vendor_locations.json')
    print('Copied to ui/vendor_locations.json')
