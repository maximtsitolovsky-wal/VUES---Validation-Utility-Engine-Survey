import openpyxl
from pathlib import Path
from datetime import datetime

excel_path = Path(r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\Excel\Vendor ASSIGN. 4.2.26.xlsx')

# Assignments to add
assignments = {
    'CEI': [9, 864, 1590],
    'PTSI': [2070, 2071, 2188],
    'alarmhawaii': [2308, 3883],
}

print(f'Loading: {excel_path}')
wb = openpyxl.load_workbook(excel_path)
print(f'Current sheets: {wb.sheetnames}')

# Header row for new sheets
headers = ['Store\nNumber', 'Store Address', 'City', 'State', 'Zip Code']

for vendor, sites in assignments.items():
    if vendor in wb.sheetnames:
        # Append to existing sheet
        ws = wb[vendor]
        start_row = ws.max_row + 1
        print(f'\n{vendor}: Adding {len(sites)} sites to existing sheet (row {start_row}+)')
        for i, site in enumerate(sites):
            ws.cell(row=start_row + i, column=1, value=site)
            print(f'  Added site {site}')
    else:
        # Create new sheet
        print(f'\n{vendor}: Creating new sheet with {len(sites)} sites')
        ws = wb.create_sheet(vendor)
        # Add headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        # Add sites
        for i, site in enumerate(sites, 2):
            ws.cell(row=i, column=1, value=site)
            print(f'  Added site {site}')

# Save
backup_path = excel_path.with_suffix('.backup.xlsx')
print(f'\nSaving backup to: {backup_path}')
# Don't overwrite existing backup
if not backup_path.exists():
    import shutil
    shutil.copy(excel_path, backup_path)

print(f'Saving updated file: {excel_path}')
wb.save(excel_path)
wb.close()

print('\n✅ Done! Vendor assignments updated.')
print(f'   CEI: +3 sites (9, 864, 1590)')
print(f'   PTSI: NEW sheet with 3 sites (2070, 2071, 2188)')
print(f'   alarmhawaii: NEW sheet with 2 sites (2308, 3883)')
