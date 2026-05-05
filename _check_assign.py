#!/usr/bin/env python
"""Check what's in the vendor assignment file."""
import openpyxl

path = r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\Excel\Vendor ASSIGN. 4.2.26.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)

print('Sheets in Vendor ASSIGN file:')
for sheet in wb.sheetnames:
    ws = wb[sheet]
    count = ws.max_row - 1  # subtract header
    print(f'  {sheet}: ~{count} rows')

# Save
with open('_vendor_assign_sheets.txt', 'w') as f:
    f.write('Sheets in Vendor ASSIGN file:\n')
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        count = ws.max_row - 1
        f.write(f'  {sheet}: ~{count} rows\n')
    f.write(f'\nTotal sheets: {len(wb.sheetnames)}\n')

print('\nSaved to _vendor_assign_sheets.txt')
