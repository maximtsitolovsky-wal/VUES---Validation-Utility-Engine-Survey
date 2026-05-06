#!/usr/bin/env python3
"""Verify 891 status."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

scout_file = r'C:\Users\vn59j7j\OneDrive - Walmart Inc\Documents\BaselinePrinter\ScoutSurveyLab.xlsm'
print(f"Loading {scout_file}...")
wb = openpyxl.load_workbook(scout_file, read_only=True, data_only=True)
ws = wb['Scout Map Data']

# Find 891
found = False
for i, row in enumerate(ws.iter_rows(min_row=2, max_col=1), start=2):
    val = row[0].value
    if val and str(val).strip().lstrip('0') == '891':
        print(f'Found 891 at row {i}: raw value = {repr(val)}')
        found = True
        break

if not found:
    print('891 NOT found in Scout Map Data')
    
# Also count total
total = sum(1 for row in ws.iter_rows(min_row=2, max_col=1) if row[0].value)
print(f'Total sites: {total}')
wb.close()
